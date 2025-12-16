"""
数据库标签页控制器
支持懒加载，数据双向同步，减少数据库请求
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import ttkbootstrap as ttk_boot
from ttkbootstrap.constants import *
import threading
import json
from typing import Dict, List, Any, Optional, Callable
from datetime import datetime
from pathlib import Path

from ..core.db_connector import PostgreSQLConnector, DBConfig
from ..config import THEME_DARK
from ..utils.file_utils import (
    safe_filename, ensure_unique_name,
    write_file_with_timestamp, write_json_with_timestamp
)


class DatabaseTabController:
    """数据库标签页控制器 - 数据缓存与双向同步"""
    
    # 完整的列配置 (列ID, 显示名称, 默认宽度)
    COLUMNS_CONFIG = {
        "models": [
            ("id", "模型ID", 150),
            ("display_name", "显示名称", 150),
            ("provider_id", "提供商ID", 100),
            ("type", "类型", 60),
            ("enabled", "启用", 50),
            ("context_window_tokens", "上下文窗口", 80),
            ("pricing", "定价信息", 100),
            ("parameters", "参数配置", 100),
            ("abilities", "能力配置", 100),
            ("user_id", "用户ID", 80),
        ],
        "providers": [
            ("id", "提供商ID", 100),
            ("name", "名称", 120),
            ("enabled", "启用", 50),
            ("sort", "排序", 50),
            ("settings", "设置", 120),
            ("config", "配置", 120),
            ("user_id", "用户ID", 80),
        ],
        "agents": [
            ("id", "助手ID", 100),
            ("title", "名称", 120),
            ("slug", "标识符", 100),
            ("description", "描述", 150),
            ("avatar", "头像", 60),
            ("model", "默认模型", 100),
            ("provider", "默认提供商", 80),
            ("system_role", "系统提示词", 150),
            ("plugins", "插件列表", 80),
            ("tags", "标签", 80),
            ("chat_config", "聊天配置", 80),
            ("params", "模型参数", 80),
            ("user_id", "用户ID", 80),
            ("created_at", "创建时间", 130),
            ("updated_at", "更新时间", 130),
        ],
        "topics": [
            ("id", "主题ID", 100),
            ("title", "标题", 200),
            ("session_id", "会话ID", 100),
            ("favorite", "收藏", 50),
            ("history_summary", "历史摘要", 150),
            ("metadata", "元数据", 100),
            ("user_id", "用户ID", 80),
            ("created_at", "创建时间", 130),
            ("updated_at", "更新时间", 130),
        ],
        "messages": [
            ("id", "消息ID", 100),
            ("role", "角色", 60),
            ("content", "内容", 300),
            ("model", "模型", 100),
            ("provider", "提供商", 80),
            ("session_id", "会话ID", 100),
            ("topic_id", "主题ID", 100),
            ("parent_id", "父消息ID", 80),
            ("tools", "工具调用", 80),
            ("metadata", "元数据", 100),
            ("reasoning", "推理过程", 80),
            ("user_id", "用户ID", 80),
            ("created_at", "创建时间", 130),
            ("updated_at", "更新时间", 130),
        ],
        # 全部对话树的列配置
        "conversations": [
            ("type", "类型", 80),
            ("model", "模型", 120),
            ("count", "数量", 60),
            ("created", "创建时间", 150),
        ],
    }
    
    def __init__(self, parent, app):
        """初始化数据库标签页控制器"""
        self.parent = parent
        self.app = app
        self.connector = None
        self.db_config = None
        self.user_id = None
        
        # 数据缓存 - 所有数据存储在这里，各标签页共享
        self.cache = {
            "agents": [],           # 完整助手列表
            "agents_full": [],      # 助手完整字段
            "topics": {},           # {agent_id: [topics]}
            "default_topics": [],   # 默认对话主题
            "messages": {},         # {topic_id: [messages]}
            "models": [],           # 模型列表
            "providers": [],        # 提供商列表
        }
        
        # 排序状态 {table_type: (column, reverse)}
        self.sort_state = {}
        
        # 选中整行模式开关
        self.select_entire_row_var = tk.BooleanVar(value=True)
        
        # 当前选中的列（用于单元格模式）
        self._selected_column = {}  # {table_type: column_index}
        
        # 分批加载的offset记录
        self._batch_offset = {"topics": 0, "messages": 0}
        self._batch_data = {"topics": [], "messages": []}  # 累积的数据
        
        # 创建UI
        self._create_ui()
    
    def _create_ui(self):
        """创建UI"""
        # 顶部工具栏（所有子标签页之上）
        top_toolbar = ttk.Frame(self.parent)
        top_toolbar.pack(fill=X, pady=2)
        
        # 左侧：数据库操作按钮
        ttk.Button(
            top_toolbar,
            text="🔌 断开连接",
            command=self._on_disconnect_click,
            bootstyle="danger-outline"
        ).pack(side=LEFT, padx=2)
        
        ttk.Button(
            top_toolbar,
            text="🔄 重载全部",
            command=self._reload_all_from_db,
            bootstyle="success-outline"
        ).pack(side=LEFT, padx=2)
        
        ttk.Separator(top_toolbar, orient=VERTICAL).pack(side=LEFT, padx=10, fill=Y, pady=2)
        
        # 导出按钮组
        ttk.Button(
            top_toolbar,
            text="📥 导出CSV",
            command=self._export_current_table_csv,
            bootstyle="info-outline"
        ).pack(side=LEFT, padx=2)
        
        ttk.Button(
            top_toolbar,
            text="📊 导出Excel",
            command=self._export_current_table_excel,
            bootstyle="info-outline"
        ).pack(side=LEFT, padx=2)
        
        ttk.Button(
            top_toolbar,
            text="📦 导出全部",
            command=self._export_all_tables,
            bootstyle="success-outline"
        ).pack(side=LEFT, padx=2)
        
        ttk.Separator(top_toolbar, orient=VERTICAL).pack(side=LEFT, padx=10, fill=Y, pady=2)
        
        # 表格适配按钮
        ttk.Button(
            top_toolbar,
            text="📐 表格适配",
            command=self._auto_fit_columns,
            bootstyle="secondary-outline"
        ).pack(side=LEFT, padx=2)
        
        ttk.Separator(top_toolbar, orient=VERTICAL).pack(side=LEFT, padx=10, fill=Y, pady=2)
        
        # 选中整行开关
        ttk.Checkbutton(
            top_toolbar,
            text="选中整行",
            variable=self.select_entire_row_var,
            bootstyle="primary-round-toggle"
        ).pack(side=LEFT, padx=10)
        
        ttk.Label(top_toolbar, text="(复制时复制整行数据)", foreground="gray").pack(side=LEFT)
        
        # 右侧：连接状态标签
        self.db_status_label = ttk.Label(top_toolbar, text="未连接", foreground="gray")
        self.db_status_label.pack(side=RIGHT, padx=10)
        
        # 子标签页 Notebook
        self.notebook = ttk.Notebook(self.parent)
        self.notebook.pack(fill=BOTH, expand=YES)
        
        # 全部对话标签页（树形视图）
        self.conversations_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.conversations_frame, text="🗨️ 全部对话")
        self._create_conversations_tab()
        
        # 模型表标签页
        self.models_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.models_frame, text="🤖 模型表")
        self._create_table_tab(self.models_frame, "models")
        
        # 提供商表标签页
        self.providers_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.providers_frame, text="🏢 提供商表")
        self._create_table_tab(self.providers_frame, "providers")
        
        # 助手表标签页
        self.agents_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.agents_frame, text="🧑‍💼 助手表")
        self._create_table_tab(self.agents_frame, "agents")
        
        # 主题表标签页
        self.topics_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.topics_frame, text="📑 主题表")
        self._create_table_tab(self.topics_frame, "topics")
        
        # 消息表标签页
        self.messages_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.messages_frame, text="💬 消息表")
        self._create_table_tab(self.messages_frame, "messages")
        
        # 搜索标签页
        self.search_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.search_frame, text="🔍 搜索")
        self._create_search_tab()
    
    def _create_conversations_tab(self):
        """创建全部对话标签页"""
        # 工具栏
        toolbar = ttk.Frame(self.conversations_frame)
        toolbar.pack(fill=X, pady=5)
        
        ttk.Button(
            toolbar, text="🔄 刷新", 
            command=self._refresh_all_data,
            bootstyle="info-outline"
        ).pack(side=LEFT, padx=2)
        
        self.conv_status_label = ttk.Label(toolbar, text="未连接", foreground="gray")
        self.conv_status_label.pack(side=RIGHT, padx=5)
        
        # 树形视图
        tree_container = ttk.Frame(self.conversations_frame)
        tree_container.pack(fill=BOTH, expand=YES)
        
        # 滚动条
        y_scroll = ttk.Scrollbar(tree_container, orient=VERTICAL)
        y_scroll.pack(side=RIGHT, fill=Y)
        
        x_scroll = ttk.Scrollbar(tree_container, orient=HORIZONTAL)
        x_scroll.pack(side=BOTTOM, fill=X)
        
        # 增加列丰富度
        self.conv_tree = ttk.Treeview(
            tree_container,
            columns=("type", "model", "count", "created"),
            show="tree headings",
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set
        )
        self.conv_tree.pack(fill=BOTH, expand=YES)
        
        y_scroll.config(command=self.conv_tree.yview)
        x_scroll.config(command=self.conv_tree.xview)
        
        # 列配置 - 支持点击排序
        self.conv_tree.heading("#0", text="名称", anchor=W, 
                               command=lambda: self._sort_conv_tree("#0"))
        self.conv_tree.heading("type", text="类型",
                               command=lambda: self._sort_conv_tree("type"))
        self.conv_tree.heading("model", text="模型",
                               command=lambda: self._sort_conv_tree("model"))
        self.conv_tree.heading("count", text="数量",
                               command=lambda: self._sort_conv_tree("count"))
        self.conv_tree.heading("created", text="创建时间",
                               command=lambda: self._sort_conv_tree("created"))
        
        self.conv_tree.column("#0", width=280, minwidth=200)
        self.conv_tree.column("type", width=80, anchor=CENTER)
        self.conv_tree.column("model", width=150)
        self.conv_tree.column("count", width=60, anchor=CENTER)
        self.conv_tree.column("created", width=150)
        
        # 绑定展开事件（懒加载）
        self.conv_tree.bind("<<TreeviewOpen>>", self._on_tree_expand)
        
        # 绑定右键菜单
        self.conv_tree.bind("<Button-3>", self._show_conv_context_menu)
    
    def _sort_conv_tree(self, col: str):
        """对话树排序 - 根据当前选中层级排序"""
        # 获取当前选中的节点
        selected = self.conv_tree.focus()
        
        # 确定排序的父节点（当前选中节点的父节点或根）
        if selected:
            parent = self.conv_tree.parent(selected)
        else:
            parent = ""
        
        # 获取当前排序状态
        current_col, reverse = self.sort_state.get(f"conv_{parent}", (None, False))
        
        # 切换排序方向
        if current_col == col:
            reverse = not reverse
        else:
            reverse = False
        
        self.sort_state[f"conv_{parent}"] = (col, reverse)
        
        # 获取该层级的所有节点
        items = self.conv_tree.get_children(parent)
        if not items:
            return
        
        # 构建排序数据
        sort_data = []
        for item in items:
            if col == "#0":
                value = self.conv_tree.item(item, "text")
            else:
                value = self.conv_tree.set(item, col)
            sort_data.append((value, item))
        
        # 排序
        try:
            if col == "count":
                sort_data.sort(key=lambda x: int(x[0]) if x[0] and x[0].isdigit() else 0, reverse=reverse)
            else:
                sort_data.sort(key=lambda x: str(x[0]).lower(), reverse=reverse)
        except:
            sort_data.sort(key=lambda x: str(x[0]).lower(), reverse=reverse)
        
        # 重新排列
        for index, (_, item) in enumerate(sort_data):
            self.conv_tree.move(item, parent, index)
        
        # 更新表头箭头
        arrow = " ▼" if reverse else " ▲"
        cols = [("#0", "名称"), ("type", "类型"), ("model", "模型"), 
                ("count", "数量"), ("created", "创建时间")]
        for c, name in cols:
            if c == col:
                self.conv_tree.heading(c, text=name + arrow)
            else:
                self.conv_tree.heading(c, text=name)
    
    def _create_table_tab(self, parent, table_type: str):
        """创建表格标签页 - 简化版，只有刷新按钮"""
        # 工具栏
        toolbar = ttk.Frame(parent)
        toolbar.pack(fill=X, pady=5)
        
        ttk.Button(
            toolbar, text="🔄 刷新", 
            command=lambda: self._refresh_table_from_cache(table_type),
            bootstyle="info-outline"
        ).pack(side=LEFT, padx=2)
        
        # 主题表和消息表添加分批加载按钮
        if table_type == "topics":
            # 主题表加载选项
            ttk.Label(toolbar, text="|", foreground="gray").pack(side=LEFT, padx=2)
            ttk.Label(toolbar, text="加载:", foreground="gray").pack(side=LEFT, padx=2)
            for count in [100, 200, 500]:
                ttk.Button(
                    toolbar, text=f"{count}个",
                    command=lambda c=count: self._load_batch_data("topics", c),
                    bootstyle="info-outline"
                ).pack(side=LEFT, padx=1)
            ttk.Button(
                toolbar, text="全部", 
                command=lambda: self._load_all_data_with_progress("topics"),
                bootstyle="success-outline"
            ).pack(side=LEFT, padx=1)
        
        elif table_type == "messages":
            # 消息表加载选项
            ttk.Label(toolbar, text="|", foreground="gray").pack(side=LEFT, padx=2)
            ttk.Label(toolbar, text="加载:", foreground="gray").pack(side=LEFT, padx=2)
            for count in [100, 200, 500, 1000, 2000]:
                ttk.Button(
                    toolbar, text=f"{count}条",
                    command=lambda c=count: self._load_batch_data("messages", c),
                    bootstyle="info-outline"
                ).pack(side=LEFT, padx=1)
            ttk.Button(
                toolbar, text="全部", 
                command=lambda: self._load_all_data_with_progress("messages"),
                bootstyle="success-outline"
            ).pack(side=LEFT, padx=1)
        
        # 状态标签
        status_label = ttk.Label(toolbar, text="0 条记录", foreground="gray")
        status_label.pack(side=RIGHT, padx=5)
        setattr(self, f"{table_type}_status_label", status_label)
        
        # 表格视图
        table_container = ttk.Frame(parent)
        table_container.pack(fill=BOTH, expand=YES)
        
        # 滚动条
        y_scroll = ttk.Scrollbar(table_container, orient=VERTICAL)
        y_scroll.pack(side=RIGHT, fill=Y)
        
        x_scroll = ttk.Scrollbar(table_container, orient=HORIZONTAL)
        x_scroll.pack(side=BOTTOM, fill=X)
        
        # 根据类型设置列
        columns_config = self.COLUMNS_CONFIG.get(table_type, [("id", "ID", 100)])
        columns = tuple(col[0] for col in columns_config)
        
        tree = ttk.Treeview(
            table_container,
            columns=columns,
            show="headings",
            yscrollcommand=y_scroll.set,
            xscrollcommand=x_scroll.set
        )
        tree.pack(fill=BOTH, expand=YES)
        
        y_scroll.config(command=tree.yview)
        x_scroll.config(command=tree.xview)
        
        # 配置列和表头排序
        for col_id, col_name, col_width in columns_config:
            tree.heading(col_id, text=col_name, anchor=W,
                        command=lambda c=col_id, t=table_type: self._sort_by_column(t, c))
            tree.column(col_id, width=col_width, minwidth=50)
        
        # 绑定左键点击事件 - 记录点击的列
        tree.bind("<Button-1>", lambda e, t=table_type: self._on_table_click(e, t))
        
        # 绑定右键菜单
        tree.bind("<Button-3>", lambda e, t=table_type: self._show_table_context_menu(e, t))
        
        setattr(self, f"{table_type}_tree", tree)
    
    def _on_table_click(self, event, table_type: str):
        """表格点击事件 - 记录点击的列"""
        tree = getattr(self, f"{table_type}_tree", None)
        if not tree:
            return
        
        # 获取点击的列
        column = tree.identify_column(event.x)
        if column:
            # column 格式为 #0, #1, #2... 转换为索引
            try:
                col_index = int(column[1:])
                self._selected_column[table_type] = col_index
            except ValueError:
                self._selected_column[table_type] = 0
    
    def _sort_by_column(self, table_type: str, col: str):
        """点击表头排序"""
        tree = getattr(self, f"{table_type}_tree", None)
        if not tree:
            return
        
        current_col, reverse = self.sort_state.get(table_type, (None, False))
        
        if current_col == col:
            reverse = not reverse
        else:
            reverse = False
        
        self.sort_state[table_type] = (col, reverse)
        
        items = [(tree.set(item, col), item) for item in tree.get_children("")]
        
        try:
            items.sort(key=lambda x: float(x[0]) if x[0] and x[0] != '-' else 0, reverse=reverse)
        except (ValueError, TypeError):
            items.sort(key=lambda x: str(x[0]).lower(), reverse=reverse)
        
        for index, (_, item) in enumerate(items):
            tree.move(item, "", index)
        
        # 更新表头
        columns_config = self.COLUMNS_CONFIG.get(table_type, [])
        for col_id, col_name, _ in columns_config:
            if col_id == col:
                arrow = " ▼" if reverse else " ▲"
                tree.heading(col_id, text=col_name + arrow)
            else:
                tree.heading(col_id, text=col_name)
    
    def _create_search_tab(self):
        """创建搜索标签页"""
        search_toolbar = ttk.Frame(self.search_frame)
        search_toolbar.pack(fill=X, pady=5)
        
        ttk.Label(search_toolbar, text="搜索:").pack(side=LEFT, padx=5)
        
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(search_toolbar, textvariable=self.search_var, width=40)
        self.search_entry.pack(side=LEFT, padx=5)
        self.search_entry.bind("<Return>", lambda e: self._execute_search())
        
        ttk.Label(search_toolbar, text="范围:").pack(side=LEFT, padx=(10, 5))
        self.search_scope_var = tk.StringVar(value="messages")
        scope_combo = ttk.Combobox(
            search_toolbar, 
            textvariable=self.search_scope_var,
            values=["messages", "topics", "agents"],
            state="readonly",
            width=12
        )
        scope_combo.pack(side=LEFT, padx=5)
        
        # 仅搜索缓存 勾选框
        self.search_cache_only_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            search_toolbar, 
            text="仅搜索缓存", 
            variable=self.search_cache_only_var,
            bootstyle="info"
        ).pack(side=LEFT, padx=5)
        
        ttk.Button(
            search_toolbar, text="🔍 搜索", 
            command=self._execute_search,
            bootstyle="primary"
        ).pack(side=LEFT, padx=5)
        
        self.search_status_label = ttk.Label(search_toolbar, text="", foreground="gray")
        self.search_status_label.pack(side=RIGHT, padx=5)
        
        result_container = ttk.Frame(self.search_frame)
        result_container.pack(fill=BOTH, expand=YES)
        
        y_scroll = ttk.Scrollbar(result_container, orient=VERTICAL)
        y_scroll.pack(side=RIGHT, fill=Y)
        
        self.search_tree = ttk.Treeview(
            result_container,
            columns=("type", "id", "content", "created"),
            show="headings",
            yscrollcommand=y_scroll.set
        )
        self.search_tree.pack(fill=BOTH, expand=YES)
        
        y_scroll.config(command=self.search_tree.yview)
        
        self.search_tree.heading("type", text="类型")
        self.search_tree.heading("id", text="ID")
        self.search_tree.heading("content", text="内容")
        self.search_tree.heading("created", text="时间")
        
        self.search_tree.column("type", width=80, anchor=CENTER)
        self.search_tree.column("id", width=150)
        self.search_tree.column("content", width=400)
        self.search_tree.column("created", width=150)
        
        self.search_offset = 0
        self.search_keyword = ""
    
    # ==================== 连接管理 ====================
    
    def set_connection(self, connector: PostgreSQLConnector, config: Dict):
        """设置数据库连接"""
        self.connector = connector
        self.db_config = config
        self.user_id = config.get("user_id")
        
        # 清空缓存
        self.cache = {
            "agents": [],
            "agents_full": [],
            "topics": {},
            "default_topics": [],
            "messages": {},
            "models": [],
            "providers": [],
        }
        self.sort_state = {}
        
        # 加载所有数据
        self._load_all_data()
    
    def _load_all_data(self):
        """加载所有数据到缓存 - 完全懒加载，不加载主题"""
        if not self.connector or not self.connector.is_connected():
            self.conv_status_label.config(text="❌ 未连接")
            return
        
        self.conv_status_label.config(text="正在加载基础数据...")
        
        # 保存已全部加载的数据标志
        all_topics_loaded = self.cache.get("_all_topics_loaded", False)
        all_topics_data = self.cache.get("_all_topics_data", None)
        all_messages_loaded = self.cache.get("_all_messages_loaded", False)
        all_messages_data = self.cache.get("_all_messages_data", None)
        
        # 【修复】保存分批加载的数据
        batch_topics_data = list(self._batch_data.get("topics", []))  # 复制列表
        batch_messages_data = list(self._batch_data.get("messages", []))
        batch_topics_offset = self._batch_offset.get("topics", 0)
        batch_messages_offset = self._batch_offset.get("messages", 0)
        
        # 【修复】保存已缓存的主题和消息数据（通过懒加载获得的）
        cached_topics = dict(self.cache.get("topics", {}))  # {agent_id: [topics]}
        cached_default_topics = list(self.cache.get("default_topics", []) or [])
        cached_messages = dict(self.cache.get("messages", {}))  # {topic_id: [messages]}
        
        def load_thread():
            try:
                # 加载全部助手（不统计主题数量）
                agents = self._query_all_agents()
                self.cache["agents"] = agents
                
                # 加载助手完整字段
                agents_full = self._query_agents_full()
                self.cache["agents_full"] = agents_full
                
                # 恢复主题数据的优先级：
                # 1. 全部加载的数据 > 2. 分批加载的数据 > 3. 懒加载的缓存数据
                if all_topics_loaded and all_topics_data:
                    # 恢复全部加载的数据
                    self.cache["_all_topics_loaded"] = True
                    self.cache["_all_topics_data"] = all_topics_data
                    self._sync_topics_to_conversation_cache(all_topics_data)
                elif batch_topics_data:
                    # 【修复】恢复分批加载的数据
                    self._batch_data["topics"] = batch_topics_data
                    self._batch_offset["topics"] = batch_topics_offset
                    self._sync_topics_to_conversation_cache(batch_topics_data)
                elif cached_topics or cached_default_topics:
                    # 【修复】恢复懒加载的缓存数据
                    self.cache["topics"] = cached_topics
                    self.cache["default_topics"] = cached_default_topics if cached_default_topics else None
                    # 更新助手的 topic_count
                    for agent in self.cache["agents"]:
                        agent_id = agent.get("id")
                        if agent_id in cached_topics:
                            agent["topic_count"] = len(cached_topics[agent_id])
                else:
                    # 没有任何缓存数据，设置为 None 表示未加载
                    self.cache["default_topics"] = None
                
                # 恢复消息数据的优先级：
                # 1. 全部加载的数据 > 2. 分批加载的数据 > 3. 懒加载的缓存数据
                if all_messages_loaded and all_messages_data:
                    # 恢复全部加载的数据
                    self.cache["_all_messages_loaded"] = True
                    self.cache["_all_messages_data"] = all_messages_data
                    self._sync_messages_to_conversation_cache(all_messages_data)
                elif batch_messages_data:
                    # 【修复】恢复分批加载的数据
                    self._batch_data["messages"] = batch_messages_data
                    self._batch_offset["messages"] = batch_messages_offset
                    self._sync_messages_to_conversation_cache(batch_messages_data)
                elif cached_messages:
                    # 【修复】恢复懒加载的缓存数据
                    self.cache["messages"] = cached_messages
                    # 更新主题的 message_count
                    for agent_id, topics in self.cache.get("topics", {}).items():
                        for topic in topics:
                            topic_id = topic.get("id")
                            if topic_id in cached_messages:
                                topic["message_count"] = len(cached_messages[topic_id])
                    if self.cache.get("default_topics"):
                        for topic in self.cache["default_topics"]:
                            topic_id = topic.get("id")
                            if topic_id in cached_messages:
                                topic["message_count"] = len(cached_messages[topic_id])
                
                # 加载模型
                models = self._query_all_models()
                self.cache["models"] = models
                
                # 加载提供商
                providers = self._query_all_providers()
                self.cache["providers"] = providers
                
                # 在主线程中更新UI
                self.parent.after(0, self._update_all_ui)
                
            except Exception as e:
                self.parent.after(0, lambda: self._show_error(f"加载失败: {e}"))
        
        threading.Thread(target=load_thread, daemon=True).start()
    
    def _refresh_all_data(self):
        """刷新所有数据"""
        self._load_all_data()
    
    # ==================== 数据库查询方法 ====================
    
    def _query_all_agents(self) -> List[Dict]:
        """查询全部助手（用于对话树）- 延迟统计主题数量"""
        # 只查询助手基本信息，不统计主题数量，避免复杂JOIN导致卡顿
        query = "SELECT id, title, slug, model, provider, created_at FROM agents"
        params = []
        if self.user_id:
            query += " WHERE user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY created_at DESC"
        
        agents = self.connector.execute_query(query, tuple(params))
        
        # 为每个助手添加 topic_count 字段，初始值为 None（表示未统计）
        for agent in agents:
            agent['topic_count'] = None
        
        return agents
    
    def _query_agents_full(self) -> List[Dict]:
        """查询助手完整字段（用于助手表）"""
        query = """SELECT id, title, slug, LEFT(description, 100) as description, 
                   avatar, model, provider, LEFT(system_role, 100) as system_role,
                   plugins::text, tags::text, chat_config::text, params::text,
                   user_id, created_at, updated_at FROM agents"""
        params = []
        if self.user_id:
            query += " WHERE user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY created_at DESC"
        
        return self.connector.execute_query(query, tuple(params))
    
    def _query_default_topics(self) -> List[Dict]:
        """查询默认对话 - 延迟统计消息数量"""
        # 只查询主题基本信息，不统计消息数量，避免JOIN导致卡顿
        query = """
            SELECT t.id, t.title, t.favorite, t.created_at
            FROM topics t
            WHERE t.session_id IS NULL OR NOT EXISTS (
                SELECT 1 FROM agents_to_sessions ats WHERE ats.session_id = t.session_id
            )
        """
        params = []
        if self.user_id:
            query += " AND t.user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY t.created_at DESC"
        
        topics = self.connector.execute_query(query, tuple(params))
        
        # 为每个主题添加 message_count 字段，初始值为 None（表示未统计）
        for topic in topics:
            topic['message_count'] = None
        
        return topics
    
    def _query_all_models(self) -> List[Dict]:
        """查询全部模型"""
        query = """SELECT id, display_name, provider_id, type, enabled, 
                   context_window_tokens, pricing::text, parameters::text, 
                   abilities::text, user_id FROM ai_models"""
        params = []
        if self.user_id:
            query += " WHERE user_id = %s"
            params.append(self.user_id)
        
        return self.connector.execute_query(query, tuple(params))
    
    def _query_all_providers(self) -> List[Dict]:
        """查询全部提供商"""
        query = """SELECT id, name, enabled, sort, 
                   settings::text, config::text, user_id FROM ai_providers"""
        params = []
        if self.user_id:
            query += " WHERE user_id = %s"
            params.append(self.user_id)
        
        return self.connector.execute_query(query, tuple(params))
    
    def _query_topics_for_agent(self, agent_id: str) -> List[Dict]:
        """查询助手的所有主题 - 延迟统计消息数量"""
        # 先检查缓存
        if agent_id in self.cache["topics"]:
            return self.cache["topics"][agent_id]
        
        # 只查询主题基本信息，不统计消息数量
        query = """
            SELECT t.id, t.title, t.favorite, t.created_at
            FROM topics t
            JOIN agents_to_sessions ats ON t.session_id = ats.session_id
            WHERE ats.agent_id = %s
        """
        params = [agent_id]
        if self.user_id:
            query += " AND t.user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY t.created_at DESC"
        
        topics = self.connector.execute_query(query, tuple(params))
        
        # 为每个主题添加 message_count 字段，初始值为 None（表示未统计）
        for topic in topics:
            topic['message_count'] = None
        
        self.cache["topics"][agent_id] = topics
        return topics
    
    def _query_messages_for_topic(self, topic_id: str) -> List[Dict]:
        """查询主题的消息"""
        # 先检查缓存
        if topic_id in self.cache["messages"]:
            return self.cache["messages"][topic_id]
        
        query = """
            SELECT id, role, content, model, created_at
            FROM messages
            WHERE topic_id = %s
        """
        params = [topic_id]
        if self.user_id:
            query += " AND user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY created_at"
        
        messages = self.connector.execute_query(query, tuple(params))
        self.cache["messages"][topic_id] = messages
        return messages
    
    def _query_all_topics(self) -> List[Dict]:
        """查询全部主题（用于主题表）"""
        query = """SELECT id, title, session_id, favorite, 
                   LEFT(history_summary, 100) as history_summary, metadata::text,
                   user_id, created_at, updated_at FROM topics"""
        params = []
        if self.user_id:
            query += " WHERE user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY created_at DESC"
        
        return self.connector.execute_query(query, tuple(params))
    
    def _query_all_messages(self) -> List[Dict]:
        """查询全部消息（用于消息表）"""
        query = """SELECT id, role, LEFT(content, 200) as content, model, provider,
                   session_id, topic_id, parent_id, tools::text, metadata::text,
                   reasoning::text, user_id, created_at, updated_at FROM messages"""
        params = []
        if self.user_id:
            query += " WHERE user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY created_at DESC"
        
        return self.connector.execute_query(query, tuple(params))
    
    def _search_database(self, keyword: str, scope: str, limit: int, offset: int) -> List[Dict]:
        """搜索数据库"""
        if scope == "messages":
            query = """
                SELECT 'message' as type, id, LEFT(content, 200) as content, created_at
                FROM messages WHERE content ILIKE %s
            """
        elif scope == "topics":
            query = """
                SELECT 'topic' as type, id, title as content, created_at
                FROM topics WHERE title ILIKE %s
            """
        elif scope == "agents":
            query = """
                SELECT 'agent' as type, id, title as content, created_at
                FROM agents WHERE title ILIKE %s OR system_role ILIKE %s
            """
        else:
            return []
        
        params = [f"%{keyword}%"]
        if scope == "agents":
            params.append(f"%{keyword}%")
        
        if self.user_id:
            query += " AND user_id = %s"
            params.append(self.user_id)
        
        query += " ORDER BY created_at DESC LIMIT %s OFFSET %s"
        params.extend([limit, offset])
        
        return self.connector.execute_query(query, tuple(params))
    
    # ==================== UI更新方法 ====================
    
    def _update_all_ui(self):
        """更新所有UI"""
        self._update_conversations_tree()
        self._update_table_from_cache("agents")
        self._update_table_from_cache("models")
        self._update_table_from_cache("providers")
        
        # 更新状态栏 - 显示主题总数和消息总数
        self._update_conv_status_label()
    
    def _update_conversations_tree(self):
        """更新对话树"""
        # 清空现有数据
        for item in self.conv_tree.get_children():
            self.conv_tree.delete(item)
        
        agents = self.cache["agents"]
        default_topics = self.cache["default_topics"]
        
        # 添加"随便聊聊"默认对话节点（始终显示）
        default_node = self.conv_tree.insert(
            "", "end",
            text="💬 随便聊聊",
            values=("默认", "", "?", ""),
            tags=("default",)
        )
        self.conv_tree.set(default_node, "type", "default:chat")
        
        # 如果默认主题未加载（None），添加"加载中..."占位符
        if default_topics is None:
            self.conv_tree.insert(default_node, "end", text="加载中...")
        # 如果已加载且有主题，显示详细信息
        elif len(default_topics) > 0:
            # 更新数量显示
            self.conv_tree.set(default_node, "count", str(len(default_topics)))
            
            for topic in default_topics:
                topic_id = topic.get("id", "")
                title = topic.get("title")
                if not title or title.strip() == "":
                    title = "默认主题"
                message_count = topic.get("message_count")
                created = self._format_datetime(topic.get("created_at"))
                
                star = "⭐ " if topic.get("favorite") else ""
                
                # 如果 message_count 为 None，显示 "?"
                count_display = "?" if message_count is None else str(message_count)
                
                node_id = self.conv_tree.insert(
                    default_node, "end",
                    text=f"📑 {star}{title}",
                    values=("主题", "", count_display, created)
                )
                self.conv_tree.set(node_id, "type", f"topic:{topic_id}")
                
                # 如果未统计或有消息，添加加载占位符
                if message_count is None or message_count > 0:
                    self.conv_tree.insert(node_id, "end", text="加载中...")
        # 如果已加载但没有主题，更新数量显示为 0
        else:
            self.conv_tree.set(default_node, "count", "0")
        
        # 添加助手节点
        for agent in agents:
            agent_id = agent.get("id", "")
            title = agent.get("title") or agent.get("slug") or agent_id[:8]
            model = agent.get("model") or ""
            topic_count = agent.get("topic_count")
            created = self._format_datetime(agent.get("created_at"))
            
            # 如果 topic_count 为 None，显示 "?"，表示未统计
            count_display = "?" if topic_count is None else str(topic_count)
            
            node_id = self.conv_tree.insert(
                "", "end",
                text=f"🧑‍💼 {title}",
                values=("助手", model, count_display, created),
                tags=("agent",)
            )
            
            # 如果未统计或有主题，添加加载占位符
            if topic_count is None or topic_count > 0:
                self.conv_tree.insert(node_id, "end", text="加载中...")
            
            self.conv_tree.set(node_id, "type", f"agent:{agent_id}")
    
    def _update_table_from_cache(self, table_type: str):
        """从缓存更新表格"""
        tree = getattr(self, f"{table_type}_tree", None)
        status_label = getattr(self, f"{table_type}_status_label", None)
        
        if not tree:
            return
        
        # 清空表格
        for item in tree.get_children():
            tree.delete(item)
        
        # 获取数据
        if table_type == "agents":
            data = self.cache.get("agents_full", [])
        elif table_type == "models":
            data = self.cache.get("models", [])
        elif table_type == "providers":
            data = self.cache.get("providers", [])
        else:
            return
        
        columns_config = self.COLUMNS_CONFIG.get(table_type, [])
        columns = [col[0] for col in columns_config]
        
        for row in data:
            values = []
            for col in columns:
                val = row.get(col, "")
                
                if col.endswith("_at") or col in ["created_at", "updated_at"]:
                    val = self._format_datetime(val)
                
                if isinstance(val, bool):
                    val = "✓" if val else "✗"
                
                if isinstance(val, str) and len(val) > 100:
                    val = val[:100] + "..."
                
                values.append(str(val) if val is not None else "")
            
            tree.insert("", "end", values=values)
        
        if status_label:
            status_label.config(text=f"{len(data)} 条记录")
    
    def _refresh_table_from_cache(self, table_type: str):
        """刷新表格 - 只从缓存读取"""
        if table_type == "topics":
            # 从缓存同步主题表
            self._sync_topics_table()
        elif table_type == "messages":
            # 从缓存同步消息表
            self._sync_messages_table()
        else:
            # 其他表从缓存更新
            self._update_table_from_cache(table_type)
    
    def _load_table_data(self, table_type: str):
        """加载表格数据"""
        if not self.connector or not self.connector.is_connected():
            return
        
        def load_thread():
            try:
                if table_type == "topics":
                    data = self._query_all_topics()
                elif table_type == "messages":
                    data = self._query_all_messages()
                else:
                    return
                
                self.parent.after(0, lambda: self._update_table_data(table_type, data))
            except Exception as e:
                self.parent.after(0, lambda: self._show_error(f"加载失败: {e}"))
        
        threading.Thread(target=load_thread, daemon=True).start()
    
    def _update_table_data(self, table_type: str, data: List[Dict]):
        """更新表格数据"""
        tree = getattr(self, f"{table_type}_tree", None)
        status_label = getattr(self, f"{table_type}_status_label", None)
        
        if not tree:
            return
        
        for item in tree.get_children():
            tree.delete(item)
        
        columns_config = self.COLUMNS_CONFIG.get(table_type, [])
        columns = [col[0] for col in columns_config]
        
        for row in data:
            values = []
            for col in columns:
                val = row.get(col, "")
                
                if col.endswith("_at"):
                    val = self._format_datetime(val)
                
                if isinstance(val, bool):
                    val = "✓" if val else "✗"
                
                if isinstance(val, str) and len(val) > 100:
                    val = val[:100] + "..."
                
                values.append(str(val) if val is not None else "")
            
            tree.insert("", "end", values=values)
        
        if status_label:
            status_label.config(text=f"{len(data)} 条记录")
    
    def _on_tree_expand(self, event):
        """树节点展开事件（懒加载）"""
        node_id = self.conv_tree.focus()
        if not node_id:
            return
        
        type_info = self.conv_tree.set(node_id, "type")
        if not type_info:
            return
        
        children = self.conv_tree.get_children(node_id)
        if len(children) == 1:
            first_child = self.conv_tree.item(children[0])
            if first_child.get("text") == "加载中...":
                self._load_children_async(node_id, type_info)
    
    def _load_children_async(self, node_id: str, type_info: str):
        """异步加载子节点"""
        def load_thread():
            try:
                # 特殊处理："default:chat" 表示默认对话节点
                if type_info == "default:chat":
                    # 加载默认主题
                    default_topics = self._query_default_topics()
                    self.cache["default_topics"] = default_topics
                    self.parent.after(0, lambda: self._insert_default_topics(node_id, default_topics))
                    # 同步刷新主题表
                    self.parent.after(100, lambda: self._sync_topics_table())
                    # 更新状态栏
                    self.parent.after(0, self._update_conv_status_label)
                    return
                
                parts = type_info.split(":")
                if len(parts) < 2:
                    return
                
                node_type, item_id = parts[0], parts[1]
                
                if node_type == "agent":
                    topics = self._query_topics_for_agent(item_id)
                    self.parent.after(0, lambda: self._insert_topics(node_id, topics))
                    # 同步刷新主题表
                    self.parent.after(100, lambda: self._sync_topics_table())
                    # 更新状态栏
                    self.parent.after(0, self._update_conv_status_label)
                    
                elif node_type == "topic":
                    messages = self._query_messages_for_topic(item_id)
                    self.parent.after(0, lambda: self._insert_messages(node_id, messages))
                    # 同步刷新消息表
                    self.parent.after(100, lambda: self._sync_messages_table())
                    # 更新状态栏
                    self.parent.after(0, self._update_conv_status_label)
                    
            except Exception as e:
                self.parent.after(0, lambda: self._show_error(f"加载失败: {e}"))
        
        threading.Thread(target=load_thread, daemon=True).start()
    
    def _insert_default_topics(self, parent_id: str, topics: List[Dict]):
        """插入默认主题节点"""
        # 清除"加载中..."占位符
        for child in self.conv_tree.get_children(parent_id):
            self.conv_tree.delete(child)
        
        if not topics:
            # 没有默认主题，更新数量为 0
            self.conv_tree.set(parent_id, "count", "0")
            return
        
        # 更新父节点的计数
        self.conv_tree.set(parent_id, "count", str(len(topics)))
        
        for topic in topics:
            topic_id = topic.get("id", "")
            title = topic.get("title")
            if not title or title.strip() == "":
                title = "默认主题"
            message_count = topic.get("message_count")
            created = self._format_datetime(topic.get("created_at"))
            
            star = "⭐ " if topic.get("favorite") else ""
            
            # 如果 message_count 为 None，显示 "?"
            count_display = "?" if message_count is None else str(message_count)
            
            node_id = self.conv_tree.insert(
                parent_id, "end",
                text=f"📑 {star}{title}",
                values=("主题", "", count_display, created)
            )
            self.conv_tree.set(node_id, "type", f"topic:{topic_id}")
            
            # 如果未统计或有消息，添加加载占位符
            if message_count is None or message_count > 0:
                self.conv_tree.insert(node_id, "end", text="加载中...")
    
    def _sync_topics_table(self):
        """同步刷新主题表 - 使用缓存数据"""
        # 优先使用全部加载的数据
        if self.cache.get("_all_topics_loaded") and self.cache.get("_all_topics_data"):
            self._update_topics_table_from_cache(self.cache["_all_topics_data"])
            return
        
        all_topics = []
        for agent_id, topics in self.cache["topics"].items():
            for topic in topics:
                # 添加agent_id信息
                topic_copy = dict(topic)
                all_topics.append(topic_copy)
        
        # 加上默认主题（如果已加载）
        if self.cache["default_topics"] is not None:
            all_topics.extend(self.cache["default_topics"])
        
        # 更新主题表
        self._update_topics_table_from_cache(all_topics)
    
    def _sync_messages_table(self):
        """同步刷新消息表 - 使用缓存数据"""
        # 优先使用全部加载的数据
        if self.cache.get("_all_messages_loaded") and self.cache.get("_all_messages_data"):
            self._update_messages_table_from_cache(self.cache["_all_messages_data"])
            return
        
        all_messages = []
        for topic_id, messages in self.cache["messages"].items():
            for msg in messages:
                msg_copy = dict(msg)
                msg_copy["topic_id"] = topic_id
                all_messages.append(msg_copy)
        
        # 更新消息表
        self._update_messages_table_from_cache(all_messages)
    
    def _update_topics_table_from_cache(self, topics: List[Dict]):
        """从缓存更新主题表"""
        tree = getattr(self, "topics_tree", None)
        status_label = getattr(self, "topics_status_label", None)
        
        if not tree:
            return
        
        for item in tree.get_children():
            tree.delete(item)
        
        columns_config = self.COLUMNS_CONFIG.get("topics", [])
        columns = [col[0] for col in columns_config]
        
        for row in topics:
            values = []
            for col in columns:
                val = row.get(col, "")
                if col.endswith("_at"):
                    val = self._format_datetime(val)
                if isinstance(val, bool):
                    val = "✓" if val else "✗"
                if isinstance(val, str) and len(val) > 100:
                    val = val[:100] + "..."
                values.append(str(val) if val is not None else "")
            tree.insert("", "end", values=values)
        
        if status_label:
            status_label.config(text=f"{len(topics)} 条记录 (缓存)")
    
    def _update_messages_table_from_cache(self, messages: List[Dict]):
        """从缓存更新消息表"""
        tree = getattr(self, "messages_tree", None)
        status_label = getattr(self, "messages_status_label", None)
        
        if not tree:
            return
        
        for item in tree.get_children():
            tree.delete(item)
        
        columns_config = self.COLUMNS_CONFIG.get("messages", [])
        columns = [col[0] for col in columns_config]
        
        for row in messages:
            values = []
            for col in columns:
                val = row.get(col, "")
                if col.endswith("_at"):
                    val = self._format_datetime(val)
                if isinstance(val, bool):
                    val = "✓" if val else "✗"
                if isinstance(val, str) and len(val) > 100:
                    val = val[:100] + "..."
                values.append(str(val) if val is not None else "")
            tree.insert("", "end", values=values)
        
        if status_label:
            status_label.config(text=f"{len(messages)} 条记录 (缓存)")
    
    def _insert_topics(self, parent_id: str, topics: List[Dict]):
        """插入主题节点"""
        for child in self.conv_tree.get_children(parent_id):
            self.conv_tree.delete(child)
        
        for topic in topics:
            topic_id = topic.get("id", "")
            title = topic.get("title")
            if not title or title.strip() == "":
                title = "默认主题"
            message_count = topic.get("message_count")
            created = self._format_datetime(topic.get("created_at"))
            
            star = "⭐ " if topic.get("favorite") else ""
            
            # 如果 message_count 为 None，显示 "?"
            count_display = "?" if message_count is None else str(message_count)
            
            node_id = self.conv_tree.insert(
                parent_id, "end",
                text=f"📑 {star}{title}",
                values=("主题", "", count_display, created)
            )
            self.conv_tree.set(node_id, "type", f"topic:{topic_id}")
            
            # 如果未统计或有消息，添加加载占位符
            if message_count is None or message_count > 0:
                self.conv_tree.insert(node_id, "end", text="加载中...")
    
    def _insert_messages(self, parent_id: str, messages: List[Dict]):
        """插入消息节点"""
        for child in self.conv_tree.get_children(parent_id):
            self.conv_tree.delete(child)
        
        for msg in messages:
            msg_id = msg.get("id", "")
            role = msg.get("role", "unknown")
            content = msg.get("content", "")[:50].replace("\n", " ")
            model = msg.get("model") or ""
            created = self._format_datetime(msg.get("created_at"))
            
            role_icon = "👤" if role == "user" else "🤖" if role == "assistant" else "⚙️"
            
            node_id = self.conv_tree.insert(
                parent_id, "end",
                text=f"{role_icon} {content}...",
                values=("消息", model, "", created)
            )
            # 设置消息节点的 type 标记，包含消息ID
            self.conv_tree.set(node_id, "type", f"message:{msg_id}")
    
    # ==================== 搜索方法 ====================
    
    def _execute_search(self):
        """执行搜索"""
        keyword = self.search_var.get().strip()
        if not keyword:
            messagebox.showwarning("警告", "请输入搜索关键词")
            return
        
        # 检查是否仅搜索缓存
        cache_only = self.search_cache_only_var.get()
        
        if not cache_only:
            if not self.connector or not self.connector.is_connected():
                messagebox.showwarning("警告", "请先连接数据库")
                return
        
        self.search_keyword = keyword
        self.search_offset = 0
        
        for item in self.search_tree.get_children():
            self.search_tree.delete(item)
        
        if cache_only:
            # 从缓存搜索
            self._search_from_cache(keyword)
        else:
            # 从数据库搜索
            self._load_search_results(100)
    
    def _search_from_cache(self, keyword: str):
        """从缓存中搜索"""
        scope = self.search_scope_var.get()
        keyword_lower = keyword.lower()
        results = []
        
        if scope == "messages":
            # 搜索缓存中的消息
            # 优先使用全部加载的数据
            if self.cache.get("_all_messages_loaded") and self.cache.get("_all_messages_data"):
                for msg in self.cache["_all_messages_data"]:
                    content = msg.get("content") or ""
                    if keyword_lower in content.lower():
                        results.append({
                            "type": "message",
                            "id": msg.get("id", ""),
                            "content": content[:200],
                            "created_at": msg.get("created_at")
                        })
            else:
                # 搜索已缓存的消息
                for topic_id, messages in self.cache["messages"].items():
                    for msg in messages:
                        content = msg.get("content") or ""
                        if keyword_lower in content.lower():
                            results.append({
                                "type": "message",
                                "id": msg.get("id", ""),
                                "content": content[:200],
                                "created_at": msg.get("created_at")
                            })
        
        elif scope == "topics":
            # 搜索缓存中的主题
            # 优先使用全部加载的数据
            if self.cache.get("_all_topics_loaded") and self.cache.get("_all_topics_data"):
                for topic in self.cache["_all_topics_data"]:
                    title = topic.get("title") or ""
                    if keyword_lower in title.lower():
                        results.append({
                            "type": "topic",
                            "id": topic.get("id", ""),
                            "content": title,
                            "created_at": topic.get("created_at")
                        })
            else:
                # 搜索已缓存的主题
                for agent_id, topics in self.cache["topics"].items():
                    for topic in topics:
                        title = topic.get("title") or ""
                        if keyword_lower in title.lower():
                            results.append({
                                "type": "topic",
                                "id": topic.get("id", ""),
                                "content": title,
                                "created_at": topic.get("created_at")
                            })
                # 搜索默认主题
                if self.cache["default_topics"]:
                    for topic in self.cache["default_topics"]:
                        title = topic.get("title") or ""
                        if keyword_lower in title.lower():
                            results.append({
                                "type": "topic",
                                "id": topic.get("id", ""),
                                "content": title,
                                "created_at": topic.get("created_at")
                            })
        
        elif scope == "agents":
            # 搜索缓存中的助手
            for agent in self.cache.get("agents_full", []):
                title = agent.get("title") or ""
                system_role = agent.get("system_role") or ""
                if keyword_lower in title.lower() or keyword_lower in system_role.lower():
                    results.append({
                        "type": "agent",
                        "id": agent.get("id", ""),
                        "content": title,
                        "created_at": agent.get("created_at")
                    })
        
        # 更新搜索结果
        self._update_search_results(results)
        
        # 更新状态
        cache_info = ""
        if scope == "messages":
            if self.cache.get("_all_messages_loaded"):
                cache_info = "(全部数据)"
            else:
                msg_count = sum(len(msgs) for msgs in self.cache["messages"].values())
                cache_info = f"(缓存{msg_count}条)"
        elif scope == "topics":
            if self.cache.get("_all_topics_loaded"):
                cache_info = "(全部数据)"
            else:
                topic_count = sum(len(topics) for topics in self.cache["topics"].values())
                if self.cache["default_topics"]:
                    topic_count += len(self.cache["default_topics"])
                cache_info = f"(缓存{topic_count}个)"
        elif scope == "agents":
            cache_info = f"(共{len(self.cache.get('agents_full', []))}个)"
        
        self.search_status_label.config(text=f"找到 {len(results)} 条结果 {cache_info}")
    
    def _load_search_results(self, count: int):
        """加载搜索结果 - 从数据库搜索"""
        if not self.search_keyword:
            return
        
        scope = self.search_scope_var.get()
        
        def search_thread():
            try:
                results = self._search_database(
                    self.search_keyword, scope, count, self.search_offset
                )
                self.search_offset += len(results)
                self.parent.after(0, lambda: self._update_search_results(results))
            except Exception as e:
                self.parent.after(0, lambda: self._show_error(f"搜索失败: {e}"))
        
        threading.Thread(target=search_thread, daemon=True).start()
    
    def _update_search_results(self, results: List[Dict]):
        """更新搜索结果"""
        for row in results:
            values = (
                row.get("type", ""),
                row.get("id", "")[:20],
                (row.get("content") or "")[:100],
                self._format_datetime(row.get("created_at"))
            )
            self.search_tree.insert("", "end", values=values)
        
        total = len(self.search_tree.get_children())
        self.search_status_label.config(text=f"找到 {total} 条结果")
    
    # ==================== 右键菜单 ====================
    
    def _show_conv_context_menu(self, event):
        """显示对话树右键菜单 - 增强版"""
        # 获取点击的项目
        item = self.conv_tree.identify_row(event.y)
        if not item:
            # 空白区域右键 - 只显示刷新
            menu = tk.Menu(self.parent, tearoff=0)
            menu.add_command(label="🔄 刷新全部", command=self._refresh_all_data)
            menu.post(event.x_root, event.y_root)
            return
        
        # 如果点击的项目不在选中列表中，则只选中该项
        current_selection = self.conv_tree.selection()
        if item not in current_selection:
            self.conv_tree.selection_set(item)
            current_selection = (item,)
        
        # 分析选中的层级类型
        has_agent = False
        has_topic = False
        has_message = False
        has_default = False
        
        for sel_item in current_selection:
            type_info = self.conv_tree.set(sel_item, "type")
            if type_info:
                if type_info == "default:chat":
                    has_default = True
                    has_agent = True
                elif type_info.startswith("agent:"):
                    has_agent = True
                elif type_info.startswith("topic:"):
                    has_topic = True
                else:
                    # 消息节点没有type标记
                    values = self.conv_tree.item(sel_item, "values")
                    if values and values[0] == "消息":
                        has_message = True
        
        # 创建动态菜单
        menu = tk.Menu(self.parent, tearoff=0)
        
        # 【新增】重载选项 - 根据选中类型显示
        reload_count = len(current_selection)
        if has_agent or has_topic or has_default:
            reload_label = f"🔃 重载选中项 ({reload_count}个)" if reload_count > 1 else "🔃 重载选中项"
            menu.add_command(label=reload_label, command=self._reload_selected_items)
            menu.add_separator()
        
        # 根据选中层级添加分割导出选项
        if has_agent:
            menu.add_command(label="📁 按助手分割导出JSON", command=self._conv_split_by_agent_json)
            menu.add_command(label="📁 按助手分割导出Markdown", command=self._conv_split_by_agent_md)
            menu.add_separator()
            menu.add_command(label="📁 按主题分割导出JSON", command=self._conv_split_by_topic_json)
            menu.add_command(label="📁 按主题分割导出Markdown", command=self._conv_split_by_topic_md)
            menu.add_separator()
            menu.add_command(label="📁 按消息分割导出JSON", command=self._conv_split_by_message_json)
            menu.add_command(label="📁 按消息分割导出Markdown", command=self._conv_split_by_message_md)
        elif has_topic:
            menu.add_command(label="📁 按主题分割导出JSON", command=self._conv_split_by_topic_json)
            menu.add_command(label="📁 按主题分割导出Markdown", command=self._conv_split_by_topic_md)
            menu.add_separator()
            menu.add_command(label="📁 按消息分割导出JSON", command=self._conv_split_by_message_json)
            menu.add_command(label="📁 按消息分割导出Markdown", command=self._conv_split_by_message_md)
        elif has_message:
            menu.add_command(label="📁 按消息分割导出JSON", command=self._conv_split_by_message_json)
            menu.add_command(label="📁 按消息分割导出Markdown", command=self._conv_split_by_message_md)
        
        # 添加分隔线
        if has_agent or has_topic or has_message:
            menu.add_separator()
        
        # 复制功能
        menu.add_command(label="📋 复制JSON到剪贴板", command=self._conv_copy_json)
        menu.add_command(label="📋 复制Markdown到剪贴板", command=self._conv_copy_md)
        menu.add_command(label="📋 复制消息内容到剪贴板", command=self._conv_copy_message_content)
        
        menu.add_separator()
        menu.add_command(label="ℹ️ 查看选中统计", command=self._conv_show_stats)
        menu.add_command(label="🔄 刷新全部", command=self._refresh_all_data)
        
        menu.post(event.x_root, event.y_root)
    
    def _show_table_context_menu(self, event, table_type: str):
        """显示表格右键菜单 - 增强版"""
        tree = getattr(self, f"{table_type}_tree", None)
        if not tree:
            return
        
        # 获取点击的项目
        item = tree.identify_row(event.y)
        if item:
            # 如果点击的项目不在选中列表中，则只选中该项
            current_selection = tree.selection()
            if item not in current_selection:
                tree.selection_set(item)
        
        current_selection = tree.selection()
        
        menu = tk.Menu(self.parent, tearoff=0)
        menu.add_command(label="🔄 刷新缓存", command=lambda: self._refresh_table_from_cache(table_type))
        
        # 【新增】重载选项 - 根据选中行从数据库重新加载
        if current_selection and table_type in ["topics", "messages", "agents"]:
            reload_count = len(current_selection)
            reload_label = f"🔃 重载选中项 ({reload_count}个)" if reload_count > 1 else "🔃 重载选中项"
            menu.add_command(label=reload_label, 
                           command=lambda: self._reload_table_selected(table_type))
        
        menu.add_separator()
        menu.add_command(label="📋 复制选中", command=lambda: self._copy_selected(table_type))
        menu.add_command(label="📋 复制全部数据", command=lambda: self._copy_all_data(table_type))
        
        # 助手表特有功能
        if table_type == "agents":
            menu.add_separator()
            menu.add_command(label="📋 复制助手提示词", command=self._copy_agent_prompt)
        
        menu.post(event.x_root, event.y_root)
    
    def _copy_selected(self, table_type: str):
        """复制选中（根据开关决定复制整行或单格）"""
        tree = getattr(self, f"{table_type}_tree", None)
        if not tree:
            return
        
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请先选择数据")
            return
        
        if self.select_entire_row_var.get():
            # 整行模式：复制选中行的所有列
            lines = []
            for item in selected:
                values = tree.item(item, "values")
                lines.append("\t".join(str(v) for v in values))
            text = "\n".join(lines)
        else:
            # 单格模式：使用记录的列索引
            col_index = self._selected_column.get(table_type, 1) - 1  # #1对应索引0
            if col_index < 0:
                col_index = 0
            
            lines = []
            for item in selected:
                values = tree.item(item, "values")
                if values and len(values) > col_index:
                    lines.append(str(values[col_index]))
                elif values:
                    lines.append(str(values[0]))
            text = "\n".join(lines)
        
        self.parent.clipboard_clear()
        self.parent.clipboard_append(text)
        
        mode_text = "整行" if self.select_entire_row_var.get() else "单元格"
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"已复制{len(selected)}条{mode_text}数据到剪贴板")
    
    def _copy_all_data(self, table_type: str):
        """复制全部数据"""
        tree = getattr(self, f"{table_type}_tree", None)
        if not tree:
            return
        
        # 获取列标题
        columns_config = self.COLUMNS_CONFIG.get(table_type, [])
        headers = [col[1] for col in columns_config]
        
        lines = ["\t".join(headers)]
        
        for item in tree.get_children():
            values = tree.item(item, "values")
            lines.append("\t".join(str(v) for v in values))
        
        text = "\n".join(lines)
        
        self.parent.clipboard_clear()
        self.parent.clipboard_append(text)
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"已复制 {len(tree.get_children())} 条数据到剪贴板")
    
    def _copy_agent_prompt(self):
        """复制助手提示词 - 从数据库现读完整数据"""
        tree = getattr(self, "agents_tree", None)
        if not tree:
            return
        
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请先选择一个助手")
            return
        
        # 收集选中的助手ID
        agent_ids = []
        for item in selected:
            values = tree.item(item, "values")
            if values:
                agent_id = values[0]  # id 列
                if agent_id:
                    agent_ids.append(agent_id)
        
        if not agent_ids:
            messagebox.showwarning("警告", "无法获取助手ID")
            return
        
        # 从数据库现读完整的提示词
        prompts = []
        if self.connector and self.connector.is_connected():
            for agent_id in agent_ids:
                full_prompt = self._query_agent_full_prompt(agent_id)
                if full_prompt:
                    prompts.append(full_prompt)
        
        if prompts:
            text = "\n\n---\n\n".join(prompts)
            self.parent.clipboard_clear()
            self.parent.clipboard_append(text)
            if self.app and hasattr(self.app, 'log_message'):
                self.app.log_message(f"✅ 已复制{len(prompts)}个助手提示词到剪贴板（完整数据）", "SUCCESS")
        else:
            messagebox.showinfo("提示", "选中的助手没有系统提示词")
    
    def _query_agent_full_prompt(self, agent_id: str) -> Optional[str]:
        """从数据库查询助手的完整提示词"""
        if not self.connector or not self.connector.is_connected():
            return None
        
        query = "SELECT system_role FROM agents WHERE id = %s"
        params = [agent_id]
        if self.user_id:
            query += " AND user_id = %s"
            params.append(self.user_id)
        
        result = self.connector.execute_query(query, tuple(params))
        if result and result[0].get("system_role"):
            return result[0]["system_role"]
        return None
    
    # ==================== 对话树导出功能 - 从数据库现读完整数据 ====================
    
    def _query_full_agent_data(self, agent_id: str) -> Dict:
        """从数据库查询助手的完整数据"""
        if not self.connector or not self.connector.is_connected():
            return {}
        
        query = "SELECT * FROM agents WHERE id = %s"
        params = [agent_id]
        if self.user_id:
            query += " AND user_id = %s"
            params.append(self.user_id)
        
        result = self.connector.execute_query(query, tuple(params))
        return result[0] if result else {}
    
    def _query_full_topics_for_agent(self, agent_id: str) -> List[Dict]:
        """从数据库查询助手的所有主题（完整数据）"""
        if not self.connector or not self.connector.is_connected():
            return []
        
        query = """
            SELECT t.* FROM topics t
            JOIN agents_to_sessions ats ON t.session_id = ats.session_id
            WHERE ats.agent_id = %s
        """
        params = [agent_id]
        if self.user_id:
            query += " AND t.user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY t.created_at DESC"
        
        return self.connector.execute_query(query, tuple(params))
    
    def _query_full_default_topics(self) -> List[Dict]:
        """从数据库查询默认对话的所有主题（完整数据）"""
        if not self.connector or not self.connector.is_connected():
            return []
        
        query = """
            SELECT t.* FROM topics t
            WHERE t.session_id IS NULL OR NOT EXISTS (
                SELECT 1 FROM agents_to_sessions ats WHERE ats.session_id = t.session_id
            )
        """
        params = []
        if self.user_id:
            query += " AND t.user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY t.created_at DESC"
        
        return self.connector.execute_query(query, tuple(params))
    
    def _query_full_messages_for_topic(self, topic_id: str) -> List[Dict]:
        """从数据库查询主题的所有消息（完整数据，不截断）"""
        if not self.connector or not self.connector.is_connected():
            return []
        
        query = "SELECT * FROM messages WHERE topic_id = %s"
        params = [topic_id]
        if self.user_id:
            query += " AND user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY created_at"
        
        return self.connector.execute_query(query, tuple(params))
    
    def _query_full_topic_data(self, topic_id: str) -> Dict:
        """从数据库查询主题的完整数据"""
        if not self.connector or not self.connector.is_connected():
            return {}
        
        query = "SELECT * FROM topics WHERE id = %s"
        params = [topic_id]
        if self.user_id:
            query += " AND user_id = %s"
            params.append(self.user_id)
        
        result = self.connector.execute_query(query, tuple(params))
        return result[0] if result else {}
    
    def _query_full_message_by_id(self, msg_id: str) -> Optional[Dict]:
        """从数据库查询单条消息的完整数据（不截断）"""
        if not self.connector or not self.connector.is_connected():
            return None
        
        query = "SELECT * FROM messages WHERE id = %s"
        params = [msg_id]
        if self.user_id:
            query += " AND user_id = %s"
            params.append(self.user_id)
        
        result = self.connector.execute_query(query, tuple(params))
        return result[0] if result else None
    
    def _get_selected_ids(self):
        """获取选中的节点ID列表（按类型分类）"""
        selection = self.conv_tree.selection()
        if not selection:
            return {"agents": [], "topics": [], "messages": [], "default": False}
        
        agent_ids = []
        topic_ids = []
        message_ids = []
        has_default = False
        
        for item in selection:
            type_info = self.conv_tree.set(item, "type")
            if not type_info:
                continue
            
            if type_info.startswith("agent:"):
                agent_id = type_info.split(":")[1]
                if agent_id not in agent_ids:
                    agent_ids.append(agent_id)
            elif type_info == "default:chat":
                has_default = True
            elif type_info.startswith("topic:"):
                topic_id = type_info.split(":")[1]
                if topic_id not in topic_ids:
                    topic_ids.append(topic_id)
            elif type_info.startswith("message:"):
                msg_id = type_info.split(":")[1]
                if msg_id not in message_ids:
                    message_ids.append(msg_id)
        
        return {"agents": agent_ids, "topics": topic_ids, "messages": message_ids, "default": has_default}
    
    def _get_selected_conv_data(self):
        """获取选中的对话数据 - 从数据库现读完整数据"""
        if not self.connector or not self.connector.is_connected():
            messagebox.showwarning("警告", "请先连接数据库")
            return None
        
        ids = self._get_selected_ids()
        if not ids["agents"] and not ids["topics"] and not ids["messages"] and not ids["default"]:
            return None
        
        all_agents = []
        all_topics = []
        all_messages = []
        
        # 处理选中的助手
        for agent_id in ids["agents"]:
            # 从数据库读取助手完整数据
            agent = self._query_full_agent_data(agent_id)
            if agent:
                all_agents.append(agent)
            
            # 从数据库读取该助手的所有主题
            topics = self._query_full_topics_for_agent(agent_id)
            for topic in topics:
                all_topics.append(topic)
                # 从数据库读取该主题的所有消息
                topic_id = topic.get("id")
                messages = self._query_full_messages_for_topic(topic_id)
                all_messages.extend(messages)
        
        # 处理默认对话
        if ids["default"]:
            topics = self._query_full_default_topics()
            for topic in topics:
                all_topics.append(topic)
                topic_id = topic.get("id")
                messages = self._query_full_messages_for_topic(topic_id)
                all_messages.extend(messages)
        
        # 处理单独选中的主题
        for topic_id in ids["topics"]:
            # 检查是否已经添加
            if any(t.get("id") == topic_id for t in all_topics):
                continue
            
            topic = self._query_full_topic_data(topic_id)
            if topic:
                all_topics.append(topic)
                messages = self._query_full_messages_for_topic(topic_id)
                all_messages.extend(messages)
        
        # 处理单独选中的消息
        for msg_id in ids["messages"]:
            # 检查是否已经添加
            if any(m.get("id") == msg_id for m in all_messages):
                continue
            
            # 从缓存或数据库获取消息
            msg = self._query_full_message_by_id(msg_id)
            if msg:
                all_messages.append(msg)
        
        return {
            "agents": all_agents,
            "topics": all_topics,
            "messages": all_messages,
            "stats": {
                "agentCount": len(all_agents),
                "topicCount": len(all_topics),
                "messageCount": len(all_messages)
            }
        }
    
    def _conv_split_by_agent_json(self):
        """按助手分割导出JSON - 从数据库现读完整数据"""
        data = self._get_selected_conv_data()
        if not data or not data["agents"]:
            messagebox.showinfo("提示", "请先选择包含助手的数据")
            return
        
        output_dir = filedialog.askdirectory(title="选择导出目录")
        if not output_dir:
            return
        
        export_dir = Path(output_dir) / f"db_agents_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        export_dir.mkdir(exist_ok=True)
        
        file_count = 0
        used_names = set()
        
        # 为每个助手收集其主题和消息（使用已从数据库读取的完整数据）
        for agent in data["agents"]:
            agent_id = agent.get("id", "")
            agent_title = agent.get("title") or agent.get("slug") or agent_id[:8]
            
            # 从已查询的数据中筛选该助手的主题
            agent_topics = [t for t in data["topics"] 
                          if self._is_topic_belong_to_agent(t, agent_id)]
            
            # 从已查询的数据中筛选该助手的消息
            topic_ids = {t.get("id") for t in agent_topics}
            agent_messages = [m for m in data["messages"] 
                            if m.get("topic_id") in topic_ids]
            
            filename = safe_filename(agent_title, agent_id)
            filename = ensure_unique_name(filename, used_names)
            
            agent_data = {
                "agent": agent,
                "topics": agent_topics,
                "messages": agent_messages
            }
            
            # 获取时间范围（数据库字段是蛇形命名）
            created_at = agent.get("created_at")
            modified_at = agent.get("updated_at") or created_at
            
            # 从消息中获取最新的修改时间
            if agent_messages:
                for msg in agent_messages:
                    msg_updated = msg.get("updated_at") or msg.get("created_at")
                    if msg_updated:
                        if not modified_at or msg_updated > modified_at:
                            modified_at = msg_updated
            
            file_path = str(export_dir / f"{filename}.json")
            write_json_with_timestamp(file_path, agent_data, created_at, modified_at)
            file_count += 1
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"✅ 按助手分割导出: {file_count}个JSON文件（完整数据）", "SUCCESS")
        messagebox.showinfo("导出成功", f"已导出{file_count}个JSON文件到:\n{export_dir}")
    
    def _is_topic_belong_to_agent(self, topic: Dict, agent_id: str) -> bool:
        """检查主题是否属于指定助手"""
        # 查询 session_id -> agent_id 映射
        session_id = topic.get("session_id")
        if not session_id:
            return False
        
        # 查数据库获取映射关系
        if not self.connector or not self.connector.is_connected():
            return False
        
        query = "SELECT 1 FROM agents_to_sessions WHERE session_id = %s AND agent_id = %s"
        result = self.connector.execute_query(query, (session_id, agent_id))
        return len(result) > 0
    
    def _conv_split_by_agent_md(self):
        """按助手分割导出Markdown - 从数据库现读完整数据"""
        data = self._get_selected_conv_data()
        if not data or not data["agents"]:
            messagebox.showinfo("提示", "请先选择包含助手的数据")
            return
        
        output_dir = filedialog.askdirectory(title="选择导出目录")
        if not output_dir:
            return
        
        export_dir = Path(output_dir) / f"db_agents_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        export_dir.mkdir(exist_ok=True)
        
        file_count = 0
        used_names = set()
        
        for agent in data["agents"]:
            agent_id = agent.get("id", "")
            agent_title = agent.get("title") or agent.get("slug") or agent_id[:8]
            
            # 从已查询的数据中筛选该助手的主题
            agent_topics = [t for t in data["topics"] 
                          if self._is_topic_belong_to_agent(t, agent_id)]
            
            # 构建主题ID到消息的映射
            topic_ids = {t.get("id") for t in agent_topics}
            messages_by_topic = {}
            for msg in data["messages"]:
                tid = msg.get("topic_id")
                if tid in topic_ids:
                    if tid not in messages_by_topic:
                        messages_by_topic[tid] = []
                    messages_by_topic[tid].append(msg)
            
            filename = safe_filename(agent_title, agent_id)
            filename = ensure_unique_name(filename, used_names)
            
            # 构建Markdown内容
            md_lines = [f"# {agent_title}", ""]
            
            for topic in agent_topics:
                topic_id = topic.get("id")
                topic_title = topic.get("title") or "未命名主题"
                md_lines.append(f"## {topic_title}")
                md_lines.append("")
                
                for msg in messages_by_topic.get(topic_id, []):
                    role = msg.get("role", "unknown")
                    content = msg.get("content", "")
                    role_icon = "👤 User" if role == "user" else "🤖 Assistant" if role == "assistant" else "⚙️ System"
                    md_lines.append(f"### {role_icon}")
                    md_lines.append("")
                    md_lines.append(content)
                    md_lines.append("")
                
                md_lines.append("---")
                md_lines.append("")
            
            # 获取时间范围（数据库字段是蛇形命名）
            agent_messages = [m for tid, msgs in messages_by_topic.items() for m in msgs if tid in topic_ids]
            created_at = agent.get("created_at")
            modified_at = agent.get("updated_at") or created_at
            
            # 从消息中获取最新的修改时间
            if agent_messages:
                for msg in agent_messages:
                    msg_updated = msg.get("updated_at") or msg.get("created_at")
                    if msg_updated:
                        if not modified_at or msg_updated > modified_at:
                            modified_at = msg_updated
            
            file_path = str(export_dir / f"{filename}.md")
            write_file_with_timestamp(file_path, "\n".join(md_lines), created_at, modified_at)
            file_count += 1
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"✅ 按助手分割导出: {file_count}个Markdown文件（完整数据）", "SUCCESS")
        messagebox.showinfo("导出成功", f"已导出{file_count}个Markdown文件到:\n{export_dir}")
    
    def _conv_split_by_topic_json(self):
        """按主题分割导出JSON - 使用已从数据库读取的完整数据"""
        data = self._get_selected_conv_data()
        if not data or not data["topics"]:
            messagebox.showinfo("提示", "请先选择包含主题的数据")
            return
        
        output_dir = filedialog.askdirectory(title="选择导出目录")
        if not output_dir:
            return
        
        export_dir = Path(output_dir) / f"db_topics_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        export_dir.mkdir(exist_ok=True)
        
        # 构建主题ID到消息的映射
        messages_by_topic = {}
        for msg in data["messages"]:
            tid = msg.get("topic_id")
            if tid:
                if tid not in messages_by_topic:
                    messages_by_topic[tid] = []
                messages_by_topic[tid].append(msg)
        
        file_count = 0
        used_names = set()
        
        for topic in data["topics"]:
            topic_id = topic.get("id", "")
            topic_title = topic.get("title") or "未命名主题"
            
            # 获取该主题的消息（从已查询的完整数据）
            messages = messages_by_topic.get(topic_id, [])
            
            filename = safe_filename(topic_title, topic_id)
            filename = ensure_unique_name(filename, used_names)
            
            topic_data = {
                "topic": topic,
                "messages": messages
            }
            
            # 获取时间信息（数据库字段是蛇形命名）
            created_at = topic.get("created_at")
            modified_at = topic.get("updated_at") or created_at
            
            # 从消息中获取最新的修改时间
            if messages:
                for msg in messages:
                    msg_updated = msg.get("updated_at") or msg.get("created_at")
                    if msg_updated:
                        if not modified_at or msg_updated > modified_at:
                            modified_at = msg_updated
            
            file_path = str(export_dir / f"{filename}.json")
            write_json_with_timestamp(file_path, topic_data, created_at, modified_at)
            file_count += 1
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"✅ 按主题分割导出: {file_count}个JSON文件（完整数据）", "SUCCESS")
        messagebox.showinfo("导出成功", f"已导出{file_count}个JSON文件到:\n{export_dir}")
    
    def _conv_split_by_topic_md(self):
        """按主题分割导出Markdown - 使用已从数据库读取的完整数据"""
        data = self._get_selected_conv_data()
        if not data or not data["topics"]:
            messagebox.showinfo("提示", "请先选择包含主题的数据")
            return
        
        output_dir = filedialog.askdirectory(title="选择导出目录")
        if not output_dir:
            return
        
        export_dir = Path(output_dir) / f"db_topics_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        export_dir.mkdir(exist_ok=True)
        
        # 构建主题ID到消息的映射
        messages_by_topic = {}
        for msg in data["messages"]:
            tid = msg.get("topic_id")
            if tid:
                if tid not in messages_by_topic:
                    messages_by_topic[tid] = []
                messages_by_topic[tid].append(msg)
        
        file_count = 0
        used_names = set()
        
        for topic in data["topics"]:
            topic_id = topic.get("id", "")
            topic_title = topic.get("title") or "未命名主题"
            
            # 获取该主题的消息（从已查询的完整数据）
            messages = messages_by_topic.get(topic_id, [])
            
            filename = safe_filename(topic_title, topic_id)
            filename = ensure_unique_name(filename, used_names)
            
            md_lines = [f"# {topic_title}", ""]
            for msg in messages:
                role = msg.get("role", "unknown")
                content = msg.get("content", "")
                role_icon = "👤 User" if role == "user" else "🤖 Assistant" if role == "assistant" else "⚙️ System"
                md_lines.append(f"## {role_icon}")
                md_lines.append("")
                md_lines.append(content)
                md_lines.append("")
            
            # 获取时间信息（数据库字段是蛇形命名）
            created_at = topic.get("created_at")
            modified_at = topic.get("updated_at") or created_at
            
            # 从消息中获取最新的修改时间
            if messages:
                for msg in messages:
                    msg_updated = msg.get("updated_at") or msg.get("created_at")
                    if msg_updated:
                        if not modified_at or msg_updated > modified_at:
                            modified_at = msg_updated
            
            file_path = str(export_dir / f"{filename}.md")
            write_file_with_timestamp(file_path, "\n".join(md_lines), created_at, modified_at)
            file_count += 1
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"✅ 按主题分割导出: {file_count}个Markdown文件（完整数据）", "SUCCESS")
        messagebox.showinfo("导出成功", f"已导出{file_count}个Markdown文件到:\n{export_dir}")
    
    def _conv_split_by_message_json(self):
        """按消息分割导出JSON"""
        data = self._get_selected_conv_data()
        if not data or not data["messages"]:
            messagebox.showinfo("提示", "请先选择包含消息的数据（需要先展开主题节点加载消息）")
            return
        
        output_dir = filedialog.askdirectory(title="选择导出目录")
        if not output_dir:
            return
        
        export_dir = Path(output_dir) / f"db_messages_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        export_dir.mkdir(exist_ok=True)
        
        file_count = 0
        used_names = set()
        
        for idx, msg in enumerate(data["messages"], 1):
            msg_id = msg.get("id", f"msg_{idx}")
            role = msg.get("role", "unknown")
            content_preview = str(msg.get("content", ""))[:30].replace("\n", " ")
            
            filename = safe_filename(f"{idx:03d}_{role}_{content_preview}", msg_id)
            filename = ensure_unique_name(filename, used_names)
            
            # 获取消息时间
            created_at = msg.get("created_at")
            modified_at = msg.get("updated_at") or created_at
            
            file_path = str(export_dir / f"{filename}.json")
            write_json_with_timestamp(file_path, msg, created_at, modified_at)
            file_count += 1
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"✅ 按消息分割导出: {file_count}个JSON文件", "SUCCESS")
        messagebox.showinfo("导出成功", f"已导出{file_count}个JSON文件到:\n{export_dir}")
    
    def _conv_split_by_message_md(self):
        """按消息分割导出Markdown"""
        data = self._get_selected_conv_data()
        if not data or not data["messages"]:
            messagebox.showinfo("提示", "请先选择包含消息的数据（需要先展开主题节点加载消息）")
            return
        
        output_dir = filedialog.askdirectory(title="选择导出目录")
        if not output_dir:
            return
        
        export_dir = Path(output_dir) / f"db_messages_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        export_dir.mkdir(exist_ok=True)
        
        file_count = 0
        used_names = set()
        
        for idx, msg in enumerate(data["messages"], 1):
            msg_id = msg.get("id", f"msg_{idx}")
            role = msg.get("role", "unknown")
            content = msg.get("content", "")
            content_preview = content[:30].replace("\n", " ")
            
            filename = safe_filename(f"{idx:03d}_{role}_{content_preview}", msg_id)
            filename = ensure_unique_name(filename, used_names)
            
            role_icon = "👤 User" if role == "user" else "🤖 Assistant" if role == "assistant" else "⚙️ System"
            md_content = f"# {role_icon}\n\n{content}\n"
            
            # 获取消息时间
            created_at = msg.get("created_at")
            modified_at = msg.get("updated_at") or created_at
            
            file_path = str(export_dir / f"{filename}.md")
            write_file_with_timestamp(file_path, md_content, created_at, modified_at)
            file_count += 1
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"✅ 按消息分割导出: {file_count}个Markdown文件", "SUCCESS")
        messagebox.showinfo("导出成功", f"已导出{file_count}个Markdown文件到:\n{export_dir}")
    
    def _conv_copy_json(self):
        """复制JSON到剪贴板"""
        data = self._get_selected_conv_data()
        if not data:
            messagebox.showinfo("提示", "请先选择数据")
            return
        
        json_str = json.dumps(data, indent=2, ensure_ascii=False, default=str)
        self.parent.clipboard_clear()
        self.parent.clipboard_append(json_str)
        
        stats = data["stats"]
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(
                f"✅ 已复制JSON到剪贴板 - "
                f"{stats['agentCount']}助手, {stats['topicCount']}主题, {stats['messageCount']}消息",
                "SUCCESS"
            )
    
    def _conv_copy_md(self):
        """复制Markdown到剪贴板"""
        data = self._get_selected_conv_data()
        if not data:
            messagebox.showinfo("提示", "请先选择数据")
            return
        
        from datetime import datetime
        md_lines = ["# 导出的对话数据", "", f"**导出时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ""]
        
        # 构建主题ID到消息的映射（使用查询到的完整数据）
        messages_by_topic = {}
        for msg in data["messages"]:
            tid = msg.get("topic_id")
            if tid:
                if tid not in messages_by_topic:
                    messages_by_topic[tid] = []
                messages_by_topic[tid].append(msg)
        
        # 如果有助手数据，按助手组织
        if data["agents"]:
            for agent in data["agents"]:
                agent_id = agent.get("id", "")
                agent_title = agent.get("title") or agent.get("slug") or agent_id[:8]
                md_lines.append(f"## 🧑‍💼 {agent_title}")
                md_lines.append("")
                
                # 筛选该助手的主题
                agent_topics = [t for t in data["topics"] 
                              if self._is_topic_belong_to_agent(t, agent_id)]
                
                for topic in agent_topics:
                    topic_id = topic.get("id")
                    topic_title = topic.get("title") or "未命名主题"
                    md_lines.append(f"### 📑 {topic_title}")
                    md_lines.append("")
                    
                    for msg in messages_by_topic.get(topic_id, []):
                        role = msg.get("role", "unknown")
                        content = msg.get("content", "")
                        role_icon = "👤 User" if role == "user" else "🤖 Assistant" if role == "assistant" else "⚙️ System"
                        md_lines.append(f"#### {role_icon}")
                        md_lines.append("")
                        md_lines.append(content)
                        md_lines.append("")
                    
                    md_lines.append("---")
                    md_lines.append("")
        
        # 如果有独立的主题数据（没有被助手包含的）
        elif data["topics"]:
            for topic in data["topics"]:
                topic_id = topic.get("id")
                topic_title = topic.get("title") or "未命名主题"
                md_lines.append(f"## 📑 {topic_title}")
                md_lines.append("")
                
                for msg in messages_by_topic.get(topic_id, []):
                    role = msg.get("role", "unknown")
                    content = msg.get("content", "")
                    role_icon = "👤 User" if role == "user" else "🤖 Assistant" if role == "assistant" else "⚙️ System"
                    md_lines.append(f"### {role_icon}")
                    md_lines.append("")
                    md_lines.append(content)
                    md_lines.append("")
                
                md_lines.append("---")
                md_lines.append("")
        
        # 如果只有消息数据（直接选中的消息）
        elif data["messages"]:
            md_lines.append("## 选中的消息")
            md_lines.append("")
            for msg in data["messages"]:
                role = msg.get("role", "unknown")
                content = msg.get("content", "")
                role_icon = "👤 User" if role == "user" else "🤖 Assistant" if role == "assistant" else "⚙️ System"
                md_lines.append(f"### {role_icon}")
                md_lines.append("")
                md_lines.append(content)
                md_lines.append("")
        
        md_text = "\n".join(md_lines)
        
        # 使用统一的剪贴板管理器
        if self.app and hasattr(self.app, 'clipboard_manager'):
            self.app.clipboard_manager.copy_to_clipboard(md_text)
        else:
            self.parent.clipboard_clear()
            self.parent.clipboard_append(md_text)
        
        stats = data["stats"]
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(
                f"✅ 已复制Markdown到剪贴板 - "
                f"{stats['topicCount']}主题, {stats['messageCount']}消息",
                "SUCCESS"
            )
    
    def _conv_copy_message_content(self):
        """复制消息内容到剪贴板（纯文本）"""
        data = self._get_selected_conv_data()
        if not data or not data["messages"]:
            messagebox.showinfo("提示", "请先选择包含消息的数据（需要先展开主题节点加载消息）")
            return
        
        content_lines = []
        for msg in data["messages"]:
            content = msg.get("content", "")
            if isinstance(content, str) and content:
                content_lines.append(content)
            elif content:
                # 如果content不是字符串（如列表或字典），尝试转换
                content_lines.append(json.dumps(content, ensure_ascii=False))
        
        if content_lines:
            combined_content = "\n\n---\n\n".join(content_lines)
            
            # 使用统一的剪贴板管理器
            if self.app and hasattr(self.app, 'clipboard_manager'):
                self.app.clipboard_manager.copy_to_clipboard(combined_content)
            else:
                self.parent.clipboard_clear()
                self.parent.clipboard_append(combined_content)
            
            if self.app and hasattr(self.app, 'log_message'):
                self.app.log_message(f"✅ 已复制{len(content_lines)}条消息内容到剪贴板", "SUCCESS")
        else:
            messagebox.showinfo("提示", "没有消息内容可复制")
    
    # ==================== 重载功能 ====================
    
    def _reload_selected_items(self):
        """
        重载选中的项目 - 从数据库重新加载数据，包含子孙节点
        支持批量选中重载
        """
        if not self.connector or not self.connector.is_connected():
            messagebox.showwarning("警告", "请先连接数据库")
            return
        
        selection = self.conv_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请先选择要重载的项目")
            return
        
        # 收集要重载的项目类型和ID
        reload_agents = []  # 需要重载的助手ID
        reload_topics = []  # 需要重载的主题ID
        reload_default = False  # 是否需要重载默认对话
        
        for item in selection:
            type_info = self.conv_tree.set(item, "type")
            if not type_info:
                continue
            
            if type_info == "default:chat":
                reload_default = True
            elif type_info.startswith("agent:"):
                agent_id = type_info.split(":")[1]
                if agent_id not in reload_agents:
                    reload_agents.append(agent_id)
            elif type_info.startswith("topic:"):
                topic_id = type_info.split(":")[1]
                if topic_id not in reload_topics:
                    reload_topics.append(topic_id)
        
        if not reload_agents and not reload_topics and not reload_default:
            messagebox.showinfo("提示", "没有可重载的项目")
            return
        
        # 显示重载进度
        total_items = len(reload_agents) + len(reload_topics) + (1 if reload_default else 0)
        self.conv_status_label.config(text=f"正在重载 {total_items} 个项目...")
        
        def reload_thread():
            try:
                reloaded_agents = 0
                reloaded_topics = 0
                reloaded_messages = 0
                
                # 重载默认对话
                if reload_default:
                    default_topics = self._query_default_topics_fresh()
                    self.cache["default_topics"] = default_topics
                    reloaded_topics += len(default_topics)
                    
                    # 重载默认对话下所有主题的消息
                    for topic in default_topics:
                        topic_id = topic.get("id")
                        messages = self._query_messages_for_topic_fresh(topic_id)
                        self.cache["messages"][topic_id] = messages
                        topic["message_count"] = len(messages)
                        reloaded_messages += len(messages)
                
                # 重载助手（包含其所有主题和消息）
                for agent_id in reload_agents:
                    # 重新查询该助手的主题
                    topics = self._query_topics_for_agent_fresh(agent_id)
                    self.cache["topics"][agent_id] = topics
                    reloaded_agents += 1
                    reloaded_topics += len(topics)
                    
                    # 更新助手的 topic_count
                    for agent in self.cache["agents"]:
                        if agent.get("id") == agent_id:
                            agent["topic_count"] = len(topics)
                            break
                    
                    # 重载该助手下所有主题的消息
                    for topic in topics:
                        topic_id = topic.get("id")
                        messages = self._query_messages_for_topic_fresh(topic_id)
                        self.cache["messages"][topic_id] = messages
                        topic["message_count"] = len(messages)
                        reloaded_messages += len(messages)
                
                # 重载单独选中的主题（包含其消息）
                for topic_id in reload_topics:
                    messages = self._query_messages_for_topic_fresh(topic_id)
                    self.cache["messages"][topic_id] = messages
                    reloaded_messages += len(messages)
                    
                    # 更新主题的 message_count
                    # 在 agent_topics 中查找
                    for agent_id, topics in self.cache["topics"].items():
                        for topic in topics:
                            if topic.get("id") == topic_id:
                                topic["message_count"] = len(messages)
                                break
                    
                    # 在 default_topics 中查找
                    if self.cache["default_topics"]:
                        for topic in self.cache["default_topics"]:
                            if topic.get("id") == topic_id:
                                topic["message_count"] = len(messages)
                                break
                
                # 在主线程中更新UI
                def update_ui():
                    self._update_conversations_tree()
                    self._sync_topics_table()
                    self._sync_messages_table()
                    self._update_conv_status_label()
                    
                    if self.app and hasattr(self.app, 'log_message'):
                        self.app.log_message(
                            f"✅ 重载完成: {reloaded_agents}个助手, "
                            f"{reloaded_topics}个主题, {reloaded_messages}条消息",
                            "SUCCESS"
                        )
                
                self.parent.after(0, update_ui)
                
            except Exception as e:
                self.parent.after(0, lambda: self._show_error(f"重载失败: {e}"))
        
        threading.Thread(target=reload_thread, daemon=True).start()
    
    def _query_default_topics_fresh(self) -> List[Dict]:
        """从数据库新鲜查询默认对话主题（不使用缓存）"""
        query = """
            SELECT t.id, t.title, t.favorite, t.created_at
            FROM topics t
            WHERE t.session_id IS NULL OR NOT EXISTS (
                SELECT 1 FROM agents_to_sessions ats WHERE ats.session_id = t.session_id
            )
        """
        params = []
        if self.user_id:
            query += " AND t.user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY t.created_at DESC"
        
        topics = self.connector.execute_query(query, tuple(params))
        
        for topic in topics:
            topic['message_count'] = None
        
        return topics
    
    def _query_topics_for_agent_fresh(self, agent_id: str) -> List[Dict]:
        """从数据库新鲜查询助手的主题（不使用缓存）"""
        query = """
            SELECT t.id, t.title, t.favorite, t.created_at
            FROM topics t
            JOIN agents_to_sessions ats ON t.session_id = ats.session_id
            WHERE ats.agent_id = %s
        """
        params = [agent_id]
        if self.user_id:
            query += " AND t.user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY t.created_at DESC"
        
        topics = self.connector.execute_query(query, tuple(params))
        
        for topic in topics:
            topic['message_count'] = None
        
        return topics
    
    def _query_messages_for_topic_fresh(self, topic_id: str) -> List[Dict]:
        """从数据库新鲜查询主题的消息（不使用缓存）"""
        query = """
            SELECT id, role, content, model, created_at
            FROM messages
            WHERE topic_id = %s
        """
        params = [topic_id]
        if self.user_id:
            query += " AND user_id = %s"
            params.append(self.user_id)
        query += " ORDER BY created_at"
        
        return self.connector.execute_query(query, tuple(params))
    
    def _reload_table_selected(self, table_type: str):
        """
        重载表格中选中的行 - 从数据库重新加载数据
        
        Args:
            table_type: 表类型 ("topics", "messages", "agents")
        """
        if not self.connector or not self.connector.is_connected():
            messagebox.showwarning("警告", "请先连接数据库")
            return
        
        tree = getattr(self, f"{table_type}_tree", None)
        if not tree:
            return
        
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请先选择要重载的数据")
            return
        
        # 收集选中行的ID
        selected_ids = []
        for item in selected:
            values = tree.item(item, "values")
            if values:
                item_id = values[0]  # 第一列是ID
                if item_id:
                    selected_ids.append(item_id)
        
        if not selected_ids:
            messagebox.showwarning("警告", "无法获取选中项的ID")
            return
        
        type_name = {"topics": "主题", "messages": "消息", "agents": "助手"}.get(table_type, "数据")
        
        def reload_thread():
            try:
                reloaded_count = 0
                
                if table_type == "topics":
                    # 重载选中的主题及其消息
                    for topic_id in selected_ids:
                        messages = self._query_messages_for_topic_fresh(topic_id)
                        self.cache["messages"][topic_id] = messages
                        
                        # 更新主题的 message_count
                        for agent_id, topics in self.cache["topics"].items():
                            for topic in topics:
                                if topic.get("id") == topic_id:
                                    topic["message_count"] = len(messages)
                                    break
                        if self.cache["default_topics"]:
                            for topic in self.cache["default_topics"]:
                                if topic.get("id") == topic_id:
                                    topic["message_count"] = len(messages)
                                    break
                        reloaded_count += 1
                
                elif table_type == "messages":
                    # 重载选中的消息（单条消息重新查询）
                    for msg_id in selected_ids:
                        msg = self._query_message_by_id(msg_id)
                        if msg:
                            topic_id = msg.get("topic_id")
                            if topic_id and topic_id in self.cache["messages"]:
                                # 更新缓存中的消息
                                for i, cached_msg in enumerate(self.cache["messages"][topic_id]):
                                    if cached_msg.get("id") == msg_id:
                                        self.cache["messages"][topic_id][i] = msg
                                        break
                            reloaded_count += 1
                
                elif table_type == "agents":
                    # 重载选中的助手及其所有主题和消息
                    for agent_id in selected_ids:
                        topics = self._query_topics_for_agent_fresh(agent_id)
                        self.cache["topics"][agent_id] = topics
                        
                        # 更新助手的 topic_count
                        for agent in self.cache["agents"]:
                            if agent.get("id") == agent_id:
                                agent["topic_count"] = len(topics)
                                break
                        
                        # 重载该助手下所有主题的消息
                        for topic in topics:
                            topic_id = topic.get("id")
                            messages = self._query_messages_for_topic_fresh(topic_id)
                            self.cache["messages"][topic_id] = messages
                            topic["message_count"] = len(messages)
                        
                        reloaded_count += 1
                
                # 在主线程中更新UI
                def update_ui():
                    self._update_conversations_tree()
                    self._sync_topics_table()
                    self._sync_messages_table()
                    self._update_conv_status_label()
                    
                    if self.app and hasattr(self.app, 'log_message'):
                        self.app.log_message(
                            f"✅ 表格重载完成: {reloaded_count}个{type_name}",
                            "SUCCESS"
                        )
                
                self.parent.after(0, update_ui)
                
            except Exception as e:
                self.parent.after(0, lambda: self._show_error(f"重载失败: {e}"))
        
        threading.Thread(target=reload_thread, daemon=True).start()
    
    def _query_message_by_id(self, msg_id: str) -> Optional[Dict]:
        """根据ID查询单条消息"""
        if not self.connector or not self.connector.is_connected():
            return None
        
        query = "SELECT id, role, content, model, topic_id, created_at FROM messages WHERE id = %s"
        params = [msg_id]
        if self.user_id:
            query += " AND user_id = %s"
            params.append(self.user_id)
        
        result = self.connector.execute_query(query, tuple(params))
        return result[0] if result else None
    
    def _conv_show_stats(self):
        """显示选中统计"""
        data = self._get_selected_conv_data()
        if not data:
            messagebox.showinfo("统计信息", "没有选中任何数据")
            return
        
        stats = data["stats"]
        stats_text = f"""选中数据统计

• 助手数据: {stats['agentCount']} 个
• 主题数据: {stats['topicCount']} 个
• 消息数据: {stats['messageCount']} 条

提示：
• 需要先展开主题节点才能获取消息数据
• 选中助手节点会包含其下所有主题和消息
"""
        
        messagebox.showinfo("选中统计", stats_text)
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(
                f"查看统计 - {stats['agentCount']}助手, {stats['topicCount']}主题, {stats['messageCount']}消息",
                "INFO"
            )
    
    # ==================== 分批加载功能 ====================
    
    def _load_batch_data(self, table_type: str, count: int):
        """
        分批加载数据（追加模式，不显示进度对话框）
        
        Args:
            table_type: 表类型 ("topics" 或 "messages")
            count: 加载数量
        """
        if not self.connector or not self.connector.is_connected():
            messagebox.showwarning("警告", "请先连接数据库")
            return
        
        type_name = "主题" if table_type == "topics" else "消息"
        status_label = getattr(self, f"{table_type}_status_label", None)
        
        # 获取当前 offset
        current_offset = self._batch_offset.get(table_type, 0)
        
        if status_label:
            status_label.config(text=f"正在加载第{current_offset + 1}-{current_offset + count}条{type_name}...")
        
        def load_thread():
            try:
                if table_type == "topics":
                    data = self._query_topics_batch(current_offset, count)
                else:  # messages
                    data = self._query_messages_batch(current_offset, count)
                
                # 在主线程中更新UI
                self.parent.after(0, lambda: self._on_batch_data_loaded(table_type, data, current_offset))
                
            except Exception as e:
                self.parent.after(0, lambda: self._show_error(f"加载失败: {e}"))
        
        threading.Thread(target=load_thread, daemon=True).start()
    
    def _on_batch_data_loaded(self, table_type: str, new_data: List[Dict], offset: int):
        """分批数据加载完成回调（追加模式）"""
        type_name = "主题" if table_type == "topics" else "消息"
        
        if not new_data:
            messagebox.showinfo("提示", f"没有更多{type_name}数据了")
            return
        
        # 追加到累积数据
        self._batch_data[table_type].extend(new_data)
        
        # 更新 offset
        self._batch_offset[table_type] = offset + len(new_data)
        
        # 更新表格 - 使用累积的全部数据
        self._update_table_data(table_type, self._batch_data[table_type])
        
        # 同步到对话树缓存
        if table_type == "topics":
            self._sync_topics_to_conversation_cache(self._batch_data[table_type])
        else:  # messages
            self._sync_messages_to_conversation_cache(self._batch_data[table_type])
        
        # 更新对话树显示
        self._update_conversations_tree()
        
        # 更新状态栏
        self._update_conv_status_label()
        
        # 更新状态标签
        total_loaded = len(self._batch_data[table_type])
        status_label = getattr(self, f"{table_type}_status_label", None)
        if status_label:
            status_label.config(text=f"{total_loaded} 条记录 (部分，可继续加载)")
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"✅ 已追加加载{len(new_data)}条{type_name}数据，累计{total_loaded}条", "SUCCESS")
    
    # ==================== 全部加载功能 ====================
    
    def _load_all_data_with_progress(self, table_type: str):
        """
        带进度对话框的全部加载功能
        
        Args:
            table_type: 表类型 ("topics" 或 "messages")
        """
        if not self.connector or not self.connector.is_connected():
            messagebox.showwarning("警告", "请先连接数据库")
            return
        
        # 显示警告
        type_name = "主题" if table_type == "topics" else "消息"
        result = messagebox.askyesno(
            "确认加载",
            f"即将加载全部{type_name}数据，数据量可能较大，可能需要较长时间。\n\n"
            f"建议：\n"
            f"• 使用树形视图按需加载单个助手或主题的数据\n"
            f"• 如需全部加载，请确保网络稳定\n\n"
            f"是否继续全部加载？"
        )
        
        if not result:
            return
        
        # 在后台线程中执行加载
        def load_thread():
            from .progress_dialog import ProgressDialog
            import time
            
            try:
                # 获取总数量
                count_query = f"SELECT COUNT(*) as count FROM {table_type}"
                if self.user_id:
                    count_query += f" WHERE user_id = '{self.user_id}'"
                count_result = self.connector.execute_query(count_query)
                total = count_result[0]["count"] if count_result else 0
                
                if total == 0:
                    self.parent.after(0, lambda: messagebox.showinfo("提示", f"没有{type_name}数据"))
                    return
                
                # 创建进度对话框 - 在主线程中创建
                self.parent.after(0, lambda: self._create_progress_dialog(
                    table_type, type_name, total
                ))
                
                # 等待对话框创建完成 - 使用 time.sleep 在后台线程等待
                progress = None
                for _ in range(20):  # 最多等待2秒
                    time.sleep(0.1)
                    progress = getattr(self, f"_{table_type}_progress_dialog", None)
                    if progress:
                        break
                
                if not progress:
                    self.parent.after(0, lambda: messagebox.showerror("错误", "无法创建进度对话框"))
                    return
                
                # 分批加载
                batch_size = 100
                loaded_count = 0
                
                all_data = []
                
                for offset in range(0, total, batch_size):
                    # 检查是否取消
                    if progress.is_cancelled:
                        break
                    
                    # 暂停控制
                    progress.wait_if_paused()
                    
                    # 查询这一批数据
                    if table_type == "topics":
                        data = self._query_topics_batch(offset, batch_size)
                    else:  # messages
                        data = self._query_messages_batch(offset, batch_size)
                    
                    all_data.extend(data)
                    loaded_count += len(data)
                    
                    # 更新进度
                    progress.update_progress(loaded_count, f"已加载: {loaded_count} / {total}")
                
                # 关闭进度对话框
                self.parent.after(0, lambda: progress.close())
                
                # 更新表格并写入共享缓存
                if not progress.is_cancelled and all_data:
                    # 将数据写入共享缓存，这样刷新时不会丢失数据
                    if table_type == "topics":
                        # 写入缓存 - 使用特殊key表示全部加载的数据
                        self.cache["_all_topics_loaded"] = True
                        self.cache["_all_topics_data"] = all_data
                        
                        # 同步到对话树缓存 - 按session_id分类主题
                        self._sync_topics_to_conversation_cache(all_data)
                        
                    else:  # messages
                        # 写入缓存 - 使用特殊key表示全部加载的数据
                        self.cache["_all_messages_loaded"] = True
                        self.cache["_all_messages_data"] = all_data
                        
                        # 同步到对话树缓存 - 按topic_id分类消息
                        self._sync_messages_to_conversation_cache(all_data)
                    
                    # 在主线程中更新表格和对话树
                    def update_ui():
                        self._update_table_data(table_type, all_data)
                        self._update_conversations_tree()
                        # 更新状态栏
                        self._update_conv_status_label()
                        messagebox.showinfo("加载完成", f"成功加载 {len(all_data)} 条{type_name}数据")
                    
                    self.parent.after(0, update_ui)
                
            except Exception as e:
                if 'progress' in locals() and progress:
                    self.parent.after(0, lambda: progress.close())
                self.parent.after(0, lambda: self._show_error(f"加载失败: {e}"))
        
        threading.Thread(target=load_thread, daemon=True).start()
    
    def _create_progress_dialog(self, table_type: str, type_name: str, total: int):
        """创建进度对话框"""
        from .progress_dialog import ProgressDialog
        
        progress = ProgressDialog(
            self.parent,
            f"加载{type_name}数据",
            f"正在加载全部{type_name}数据，请稍候...\n可以暂停或取消操作。",
            total
        )
        setattr(self, f"_{table_type}_progress_dialog", progress)
    
    def _query_topics_batch(self, offset: int, limit: int) -> List[Dict]:
        """分批查询主题"""
        query = """SELECT id, title, session_id, favorite, 
                   LEFT(history_summary, 100) as history_summary, metadata::text,
                   user_id, created_at, updated_at FROM topics"""
        params = []
        if self.user_id:
            query += " WHERE user_id = %s"
            params.append(self.user_id)
        query += f" ORDER BY created_at DESC LIMIT {limit} OFFSET {offset}"
        
        return self.connector.execute_query(query, tuple(params))
    
    def _query_messages_batch(self, offset: int, limit: int) -> List[Dict]:
        """分批查询消息"""
        query = """SELECT id, role, LEFT(content, 200) as content, model, provider,
                   session_id, topic_id, parent_id, tools::text, metadata::text,
                   reasoning::text, user_id, created_at, updated_at FROM messages"""
        params = []
        if self.user_id:
            query += " WHERE user_id = %s"
            params.append(self.user_id)
        query += f" ORDER BY created_at DESC LIMIT {limit} OFFSET {offset}"
        
        return self.connector.execute_query(query, tuple(params))
    
    # ==================== 缓存同步方法 ====================
    
    def _sync_topics_to_conversation_cache(self, all_topics: List[Dict]):
        """
        将全部加载的主题数据同步到对话树缓存
        
        Args:
            all_topics: 全部主题数据列表
        """
        # 需要查询 session_id 与 agent_id 的对应关系
        if not self.connector or not self.connector.is_connected():
            return
        
        # 获取所有 session_id -> agent_id 的映射
        query = "SELECT session_id, agent_id FROM agents_to_sessions"
        mapping_result = self.connector.execute_query(query)
        
        session_to_agent = {}
        for row in mapping_result:
            session_id = row.get("session_id")
            agent_id = row.get("agent_id")
            if session_id and agent_id:
                session_to_agent[session_id] = agent_id
        
        # 按 agent_id 分组主题
        agent_topics = {}  # {agent_id: [topics]}
        default_topics = []  # 没有关联助手的主题
        
        for topic in all_topics:
            session_id = topic.get("session_id")
            
            if session_id and session_id in session_to_agent:
                agent_id = session_to_agent[session_id]
                if agent_id not in agent_topics:
                    agent_topics[agent_id] = []
                agent_topics[agent_id].append(topic)
            else:
                # 没有 session_id 或没有关联助手的主题归入默认对话
                default_topics.append(topic)
        
        # 更新缓存
        self.cache["topics"] = agent_topics
        self.cache["default_topics"] = default_topics
        
        # 更新助手的 topic_count
        for agent in self.cache["agents"]:
            agent_id = agent.get("id")
            if agent_id in agent_topics:
                agent["topic_count"] = len(agent_topics[agent_id])
            else:
                agent["topic_count"] = 0
    
    def _sync_messages_to_conversation_cache(self, all_messages: List[Dict]):
        """
        将全部加载的消息数据同步到对话树缓存
        
        Args:
            all_messages: 全部消息数据列表
        """
        # 按 topic_id 分组消息
        topic_messages = {}  # {topic_id: [messages]}
        
        for msg in all_messages:
            topic_id = msg.get("topic_id")
            if topic_id:
                if topic_id not in topic_messages:
                    topic_messages[topic_id] = []
                topic_messages[topic_id].append(msg)
        
        # 更新缓存
        self.cache["messages"] = topic_messages
        
        # 更新主题的 message_count
        # 更新 agent_topics 中的主题
        for agent_id, topics in self.cache["topics"].items():
            for topic in topics:
                topic_id = topic.get("id")
                if topic_id in topic_messages:
                    topic["message_count"] = len(topic_messages[topic_id])
                else:
                    topic["message_count"] = 0
        
        # 更新 default_topics 中的主题
        if self.cache["default_topics"]:
            for topic in self.cache["default_topics"]:
                topic_id = topic.get("id")
                if topic_id in topic_messages:
                    topic["message_count"] = len(topic_messages[topic_id])
                else:
                    topic["message_count"] = 0
    
    # ==================== 辅助方法 ====================
    
    def _update_conv_status_label(self):
        """更新对话树状态栏 - 显示主题总数和消息总数"""
        # 统计主题总数
        topic_count = 0
        topic_loaded = False
        
        # 如果全部加载过主题
        if self.cache.get("_all_topics_loaded") and self.cache.get("_all_topics_data"):
            topic_count = len(self.cache["_all_topics_data"])
            topic_loaded = True
        else:
            # 累加缓存中的主题数量
            for agent_id, topics in self.cache["topics"].items():
                topic_count += len(topics)
            if self.cache["default_topics"] is not None:
                topic_count += len(self.cache["default_topics"])
                topic_loaded = True if topic_count > 0 or len(self.cache["topics"]) > 0 else False
        
        # 统计消息总数
        message_count = 0
        message_loaded = False
        
        # 如果全部加载过消息
        if self.cache.get("_all_messages_loaded") and self.cache.get("_all_messages_data"):
            message_count = len(self.cache["_all_messages_data"])
            message_loaded = True
        else:
            # 累加缓存中的消息数量
            for topic_id, messages in self.cache["messages"].items():
                message_count += len(messages)
            message_loaded = message_count > 0
        
        # 格式化显示
        topic_display = str(topic_count) if topic_loaded else "?"
        message_display = str(message_count) if message_loaded else "?"
        
        self.conv_status_label.config(text=f"✅ {topic_display}个主题, {message_display}条消息")
    
    def _format_datetime(self, dt) -> str:
        """格式化日期时间"""
        if dt is None:
            return ""
        if isinstance(dt, datetime):
            return dt.strftime("%Y-%m-%d %H:%M")
        if isinstance(dt, str):
            try:
                if "T" in dt:
                    dt_obj = datetime.fromisoformat(dt.replace("Z", "+00:00"))
                    return dt_obj.strftime("%Y-%m-%d %H:%M")
            except:
                pass
            return dt[:16] if len(dt) > 16 else dt
        return str(dt)
    
    def _show_error(self, message: str):
        """显示错误消息"""
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(message, "ERROR")
        messagebox.showerror("错误", message)
    
    def _on_disconnect_click(self):
        """断开连接按钮点击事件"""
        if not self.connector or not self.connector.is_connected():
            messagebox.showinfo("提示", "当前未连接数据库")
            return
        
        # 确认断开
        result = messagebox.askyesno(
            "确认断开",
            "确定要断开数据库连接吗？\n\n"
            "断开后缓存数据将被清空，需要重新连接才能使用数据库功能。"
        )
        
        if not result:
            return
        
        # 断开连接
        self.disconnect()
        
        # 清空缓存
        self.cache = {
            "agents": [],
            "agents_full": [],
            "topics": {},
            "default_topics": [],
            "messages": {},
            "models": [],
            "providers": [],
        }
        self._batch_data = {"topics": [], "messages": []}
        self._batch_offset = {"topics": 0, "messages": 0}
        
        # 清空所有表格
        for item in self.conv_tree.get_children():
            self.conv_tree.delete(item)
        
        for table_type in ["models", "providers", "agents", "topics", "messages"]:
            tree = getattr(self, f"{table_type}_tree", None)
            if tree:
                for item in tree.get_children():
                    tree.delete(item)
        
        # 更新状态
        self.conv_status_label.config(text="❌ 已断开连接")
        self.db_status_label.config(text="❌ 未连接", foreground="gray")
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message("✅ 已断开数据库连接", "SUCCESS")
        
        # 通知主窗口（如果有回调）
        if self.app and hasattr(self.app, 'on_db_disconnected'):
            self.app.on_db_disconnected()
    
    def _reload_all_from_db(self):
        """
        重载全部数据 - 清空缓存，从数据库重新加载所有数据
        与刷新不同，这会清空所有缓存数据，重新从数据库加载
        """
        if not self.connector or not self.connector.is_connected():
            messagebox.showwarning("警告", "请先连接数据库")
            return
        
        # 确认重载
        result = messagebox.askyesno(
            "确认重载",
            "确定要重载全部数据吗？\n\n"
            "这将清空所有缓存数据，包括已加载的主题和消息，\n"
            "然后从数据库重新加载基础数据。\n\n"
            "主题和消息将需要重新按需加载。"
        )
        
        if not result:
            return
        
        # 清空缓存
        self.cache = {
            "agents": [],
            "agents_full": [],
            "topics": {},
            "default_topics": None,  # None 表示未加载
            "messages": {},
            "models": [],
            "providers": [],
        }
        self._batch_data = {"topics": [], "messages": []}
        self._batch_offset = {"topics": 0, "messages": 0}
        
        # 显示状态
        self.conv_status_label.config(text="正在重载全部数据...")
        self.db_status_label.config(text="正在重载...", foreground="orange")
        
        # 从数据库重新加载
        def reload_thread():
            try:
                # 加载全部助手
                agents = self._query_all_agents()
                self.cache["agents"] = agents
                
                # 加载助手完整字段
                agents_full = self._query_agents_full()
                self.cache["agents_full"] = agents_full
                
                # 加载模型
                models = self._query_all_models()
                self.cache["models"] = models
                
                # 加载提供商
                providers = self._query_all_providers()
                self.cache["providers"] = providers
                
                # 在主线程中更新UI
                def update_ui():
                    self._update_all_ui()
                    self.db_status_label.config(text="✅ 已连接", foreground="green")
                    
                    if self.app and hasattr(self.app, 'log_message'):
                        self.app.log_message(
                            f"✅ 重载完成: {len(agents)}个助手, "
                            f"{len(models)}个模型, {len(providers)}个提供商",
                            "SUCCESS"
                        )
                
                self.parent.after(0, update_ui)
                
            except Exception as e:
                self.parent.after(0, lambda: self._show_error(f"重载失败: {e}"))
                self.parent.after(0, lambda: self.db_status_label.config(text="❌ 重载失败", foreground="red"))
        
        threading.Thread(target=reload_thread, daemon=True).start()
    
    def configure_theme(self, theme: str):
        """配置主题"""
        pass
    
    def disconnect(self):
        """断开数据库连接"""
        if self.connector:
            try:
                self.connector.disconnect()
            except:
                pass
            self.connector = None
        
        self.conv_status_label.config(text="❌ 已断开连接")
    
    # ==================== 导出功能 ====================
    
    def _export_current_table_csv(self):
        """导出当前选中标签页的表格为CSV"""
        current_tab_idx = self.notebook.index(self.notebook.select())
        tab_names = ["conversations", "models", "providers", "agents", "topics", "messages", "search"]
        
        if current_tab_idx >= len(tab_names):
            messagebox.showwarning("警告", "无法识别当前标签页")
            return
        
        table_type = tab_names[current_tab_idx]
        
        if table_type == "conversations":
            messagebox.showinfo("提示", "对话树不支持导出CSV，请使用右键菜单的分割导出功能")
            return
        
        tree = getattr(self, f"{table_type}_tree", None)
        if not tree:
            messagebox.showwarning("警告", "无法获取当前表格")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"{table_type}_{timestamp}"

        file_path = filedialog.asksaveasfilename(
            title="导出CSV",
            defaultextension=".csv",
            initialfile=default_name,
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            import csv
            
            columns_config = self.COLUMNS_CONFIG.get(table_type, [])
            headers = [col[1] for col in columns_config]
            
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                
                for item in tree.get_children():
                    values = tree.item(item, "values")
                    writer.writerow(values)
            
            if self.app and hasattr(self.app, 'log_message'):
                self.app.log_message(f"✅ 已导出CSV: {file_path}", "SUCCESS")
            messagebox.showinfo("导出成功", f"已导出CSV文件到:\n{file_path}")
            
        except Exception as e:
            self._show_error(f"导出CSV失败: {e}")
    
    def _export_current_table_excel(self):
        """导出当前选中标签页的表格为Excel"""
        current_tab_idx = self.notebook.index(self.notebook.select())
        tab_names = ["conversations", "models", "providers", "agents", "topics", "messages", "search"]
        
        if current_tab_idx >= len(tab_names):
            messagebox.showwarning("警告", "无法识别当前标签页")
            return
        
        table_type = tab_names[current_tab_idx]
        
        if table_type == "conversations":
            messagebox.showinfo("提示", "对话树不支持导出Excel，请使用右键菜单的分割导出功能")
            return
        
        tree = getattr(self, f"{table_type}_tree", None)
        if not tree:
            messagebox.showwarning("警告", "无法获取当前表格")
            return
        
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("错误", "需要安装openpyxl库才能导出Excel\n请运行: pip install openpyxl")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"{table_type}_{timestamp}"

        file_path = filedialog.asksaveasfilename(
            title="导出Excel",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = table_type.capitalize()
            
            columns_config = self.COLUMNS_CONFIG.get(table_type, [])
            headers = [col[1] for col in columns_config]
            
            ws.append(headers)
            
            for item in tree.get_children():
                values = tree.item(item, "values")
                ws.append(list(values))
            
            wb.save(file_path)
            
            if self.app and hasattr(self.app, 'log_message'):
                self.app.log_message(f"✅ 已导出Excel: {file_path}", "SUCCESS")
            messagebox.showinfo("导出成功", f"已导出Excel文件到:\n{file_path}")
            
        except Exception as e:
            self._show_error(f"导出Excel失败: {e}")
    
    def _export_all_tables(self):
        """导出所有表格到一个Excel文件"""
        try:
            import openpyxl
        except ImportError:
            messagebox.showerror("错误", "需要安装openpyxl库才能导出Excel\n请运行: pip install openpyxl")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"lobechat_db_export_{timestamp}"

        file_path = filedialog.asksaveasfilename(
            title="导出全部表格",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import Workbook
            
            wb = Workbook()
            wb.remove(wb.active)
            
            table_types = ["models", "providers", "agents", "topics", "messages"]
            exported_count = 0
            
            for table_type in table_types:
                tree = getattr(self, f"{table_type}_tree", None)
                if not tree:
                    continue
                
                ws = wb.create_sheet(title=table_type.capitalize())
                
                columns_config = self.COLUMNS_CONFIG.get(table_type, [])
                headers = [col[1] for col in columns_config]
                
                ws.append(headers)
                
                row_count = 0
                for item in tree.get_children():
                    values = tree.item(item, "values")
                    ws.append(list(values))
                    row_count += 1
                
                if row_count > 0:
                    exported_count += 1
            
            if exported_count == 0:
                messagebox.showinfo("提示", "没有可导出的数据")
                return
            
            wb.save(file_path)
            
            if self.app and hasattr(self.app, 'log_message'):
                self.app.log_message(f"✅ 已导出全部表格({exported_count}个): {file_path}", "SUCCESS")
            messagebox.showinfo("导出成功", f"已导出{exported_count}个表格到:\n{file_path}")
            
        except Exception as e:
            self._show_error(f"导出全部表格失败: {e}")
            self._show_error(f"导出全部表格失败: {e}")
    
    def _auto_fit_columns(self):
        """自动适配当前表格的列宽"""
        current_tab_idx = self.notebook.index(self.notebook.select())
        tab_names = ["conversations", "models", "providers", "agents", "topics", "messages", "search"]
        
        if current_tab_idx >= len(tab_names):
            return
        
        table_type = tab_names[current_tab_idx]
        
        # 对话树使用特殊处理
        if table_type == "conversations":
            tree = self.conv_tree
            # 对话树只适配可见列
            for col in ["#0", "type", "model", "count", "created"]:
                max_width = 100
                
                if col == "#0":
                    for item in tree.get_children(""):
                        text = tree.item(item, "text")
                        width = len(str(text)) * 8
                        if width > max_width:
                            max_width = min(width, 400)
                else:
                    for item in tree.get_children(""):
                        val = tree.set(item, col)
                        width = len(str(val)) * 8
                        if width > max_width:
                            max_width = min(width, 400)
                
                tree.column(col, width=max_width)
        else:
            tree = getattr(self, f"{table_type}_tree", None)
            if not tree:
                return
            
            columns_config = self.COLUMNS_CONFIG.get(table_type, [])
            
            for col_id, col_name, _ in columns_config:
                max_width = len(col_name) * 10
                
                for item in tree.get_children():
                    val = tree.set(item, col_id)
                    width = len(str(val)) * 8
                    if width > max_width:
                        max_width = min(width, 500)
                
                tree.column(col_id, width=max(max_width, 80))
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message("✅ 已自动适配列宽", "SUCCESS")
    
    # ==================== 对话树展开/收缩功能 ====================
    
    def _expand_selected_items(self):
        """展开选中的节点（支持批量）"""
        selection = self.conv_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请先选择要展开的节点")
            return
        
        for item in selection:
            self._expand_item_recursive(item)
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"✅ 已展开{len(selection)}个节点", "SUCCESS")
    
    def _collapse_selected_items(self):
        """收缩选中的节点（支持批量）"""
        selection = self.conv_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请先选择要收缩的节点")
            return
        
        for item in selection:
            self._collapse_item_recursive(item)
        
        if self.app and hasattr(self.app, 'log_message'):
            self.app.log_message(f"✅ 已收缩{len(selection)}个节点", "SUCCESS")
    
    def _expand_item_recursive(self, item):
        """递归展开节点及其所有子节点"""
        # 先展开当前节点
        self.conv_tree.item(item, open=True)
        
        # 如果有"加载中..."占位符，触发懒加载
        children = self.conv_tree.get_children(item)
        if len(children) == 1:
            first_child = self.conv_tree.item(children[0])
            if first_child.get("text") == "加载中...":
                type_info = self.conv_tree.set(item, "type")
                if type_info:
                    self._load_children_async(item, type_info)
                    # 等待加载完成后再展开子节点
                    return
        
        # 递归展开所有子节点
        for child in children:
            self._expand_item_recursive(child)
    
    def _collapse_item_recursive(self, item):
        """递归收缩节点及其所有子节点"""
        # 先收缩所有子节点
        for child in self.conv_tree.get_children(item):
            self._collapse_item_recursive(child)
        
        # 最后收缩当前节点
        self.conv_tree.item(item, open=False)
