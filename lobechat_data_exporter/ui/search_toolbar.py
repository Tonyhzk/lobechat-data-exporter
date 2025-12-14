"""
搜索工具栏组件
提供搜索、筛选和导出功能
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from ttkbootstrap.constants import *
from typing import Dict, List, Any, Optional, Callable
import csv
import re


class MultiSelectListbox(ttk.Frame):
    """多选列表框组件（下拉式）"""
    
    def __init__(self, parent, values: List[str], on_change: Callable = None):
        super().__init__(parent)
        self.values = values
        self.on_change = on_change
        self.check_vars = {}
        self.dropdown_window = None
        
        self._create_ui()
        self._init_vars()
    
    def _create_ui(self):
        """创建UI"""
        # 显示按钮
        self.display_btn = ttk.Button(
            self,
            text="全部列 ▼",
            command=self._toggle_dropdown,
            bootstyle="outline",
            width=12
        )
        self.display_btn.pack(side=LEFT)
    
    def _init_vars(self):
        """初始化选择变量"""
        for value in self.values:
            self.check_vars[value] = tk.BooleanVar(value=(value == "全部列"))
    
    def _toggle_dropdown(self):
        """切换下拉框显示"""
        if self.dropdown_window and self.dropdown_window.winfo_exists():
            self._hide_dropdown()
        else:
            self._show_dropdown()
    
    def _show_dropdown(self):
        """显示下拉框"""
        if self.dropdown_window and self.dropdown_window.winfo_exists():
            return
        
        # 创建Toplevel窗口
        self.dropdown_window = tk.Toplevel(self.winfo_toplevel())
        self.dropdown_window.overrideredirect(True)  # 无边框
        self.dropdown_window.attributes("-topmost", True)
        
        # 计算位置 - 基于按钮的屏幕坐标
        self.display_btn.update_idletasks()
        btn_x = self.display_btn.winfo_rootx()
        btn_y = self.display_btn.winfo_rooty()
        btn_height = self.display_btn.winfo_height()
        
        # 窗口位置在按钮正下方
        self.dropdown_window.geometry(f"+{btn_x}+{btn_y + btn_height}")
        
        # 外框
        frame = ttk.Frame(self.dropdown_window, relief="solid", borderwidth=1)
        frame.pack(fill=BOTH, expand=YES)
        
        # 全选/全不选按钮
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=X, padx=5, pady=3)
        
        ttk.Button(
            btn_frame,
            text="全选",
            command=self._select_all,
            bootstyle="info-outline",
            width=6
        ).pack(side=LEFT, padx=2)
        
        ttk.Button(
            btn_frame,
            text="清除",
            command=self._deselect_all,
            bootstyle="info-outline",
            width=6
        ).pack(side=LEFT, padx=2)
        
        # 分隔线
        ttk.Separator(frame, orient=HORIZONTAL).pack(fill=X, padx=5)
        
        # 复选框列表（带滚动）
        list_frame = ttk.Frame(frame)
        list_frame.pack(fill=BOTH, expand=YES, padx=5, pady=3)
        
        # 限制高度
        canvas = tk.Canvas(list_frame, highlightthickness=0, height=150, width=150)
        scrollbar = ttk.Scrollbar(list_frame, orient=VERTICAL, command=canvas.yview)
        inner_frame = ttk.Frame(canvas)
        
        inner_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=inner_frame, anchor=NW)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 只有内容多时才显示滚动条
        if len(self.values) > 8:
            scrollbar.pack(side=RIGHT, fill=Y)
        canvas.pack(side=LEFT, fill=BOTH, expand=YES)
        
        for value in self.values:
            if value not in self.check_vars:
                self.check_vars[value] = tk.BooleanVar(value=(value == "全部列"))
            
            cb = ttk.Checkbutton(
                inner_frame,
                text=value,
                variable=self.check_vars[value],
                command=self._on_check_changed,
                bootstyle="primary"
            )
            cb.pack(anchor=W, pady=1)
        
        # 绑定点击其他地方关闭
        self.dropdown_window.bind("<FocusOut>", self._on_focus_out)
        self.winfo_toplevel().bind("<Button-1>", self._on_root_click, add="+")
    
    def _hide_dropdown(self):
        """隐藏下拉框"""
        if self.dropdown_window and self.dropdown_window.winfo_exists():
            self.dropdown_window.destroy()
        self.dropdown_window = None
        try:
            self.winfo_toplevel().unbind("<Button-1>")
        except:
            pass
    
    def _on_focus_out(self, event):
        """失去焦点时关闭"""
        # 延迟检查，避免点击下拉框内部时关闭
        self.after(100, self._check_and_hide)
    
    def _check_and_hide(self):
        """检查是否应该隐藏"""
        if self.dropdown_window and self.dropdown_window.winfo_exists():
            try:
                # 检查鼠标是否在下拉框内
                x = self.winfo_toplevel().winfo_pointerx()
                y = self.winfo_toplevel().winfo_pointery()
                
                wx = self.dropdown_window.winfo_rootx()
                wy = self.dropdown_window.winfo_rooty()
                ww = self.dropdown_window.winfo_width()
                wh = self.dropdown_window.winfo_height()
                
                if not (wx <= x <= wx + ww and wy <= y <= wy + wh):
                    self._hide_dropdown()
            except:
                pass
    
    def _on_root_click(self, event):
        """主窗口点击事件"""
        if not self.dropdown_window or not self.dropdown_window.winfo_exists():
            return
        
        # 检查是否点击在下拉框外
        try:
            x, y = event.x_root, event.y_root
            wx = self.dropdown_window.winfo_rootx()
            wy = self.dropdown_window.winfo_rooty()
            ww = self.dropdown_window.winfo_width()
            wh = self.dropdown_window.winfo_height()
            
            # 也检查是否点击在按钮上
            bx = self.display_btn.winfo_rootx()
            by = self.display_btn.winfo_rooty()
            bw = self.display_btn.winfo_width()
            bh = self.display_btn.winfo_height()
            
            if not (wx <= x <= wx + ww and wy <= y <= wy + wh) and \
               not (bx <= x <= bx + bw and by <= y <= by + bh):
                self._hide_dropdown()
        except:
            pass
    
    def _on_check_changed(self):
        """复选框状态改变"""
        selected = self.get_selected()
        
        # 更新显示文本
        if not selected or "全部列" in selected:
            self.display_btn.config(text="全部列 ▼")
        elif len(selected) == 1:
            text = selected[0][:8] + "..." if len(selected[0]) > 8 else selected[0]
            self.display_btn.config(text=f"{text} ▼")
        else:
            self.display_btn.config(text=f"{len(selected)}列 ▼")
        
        if self.on_change:
            self.on_change(selected)
    
    def _select_all(self):
        """全选"""
        for var in self.check_vars.values():
            var.set(True)
        self._on_check_changed()
    
    def _deselect_all(self):
        """全不选"""
        for var in self.check_vars.values():
            var.set(False)
        # 至少保留"全部列"
        if "全部列" in self.check_vars:
            self.check_vars["全部列"].set(True)
        self._on_check_changed()
    
    def get_selected(self) -> List[str]:
        """获取选中的列"""
        return [k for k, v in self.check_vars.items() if v.get()]
    
    def update_values(self, values: List[str]):
        """更新可选值"""
        self.values = values
        # 重置选择状态
        self.check_vars = {}
        for value in values:
            self.check_vars[value] = tk.BooleanVar(value=(value == "全部列"))
        self.display_btn.config(text="全部列 ▼")


class SearchToolbar(ttk.Frame):
    """搜索工具栏"""
    
    def __init__(self, parent, app, on_search: Callable, on_export: Callable,
                 on_search_all: Callable = None, on_prev: Callable = None, on_next: Callable = None):
        """
        初始化搜索工具栏
        
        Args:
            parent: 父容器
            app: 主应用实例
            on_search: 搜索单个回调函数（定位）
            on_export: 导出回调函数
            on_search_all: 搜索全部回调函数
            on_prev: 上一个回调函数
            on_next: 下一个回调函数
        """
        super().__init__(parent)
        self.app = app
        self.on_search_callback = on_search
        self.on_export_callback = on_export
        self.on_search_all_callback = on_search_all
        self.on_prev_callback = on_prev
        self.on_next_callback = on_next
        
        # 搜索选项
        self.search_var = tk.StringVar()
        self.search_full_text_var = tk.BooleanVar(value=False)
        
        # 可用列列表
        self.available_columns = ["全部列"]
        
        self._create_ui()
    
    def _create_ui(self):
        """创建UI"""
        # 搜索区域
        search_frame = ttk.Frame(self)
        search_frame.pack(side=LEFT, fill=X, expand=YES, padx=(0, 10))
        
        # 搜索标签
        ttk.Label(search_frame, text="🔍 搜索:").pack(side=LEFT, padx=(0, 5))
        
        # 搜索输入框
        self.search_entry = ttk.Entry(
            search_frame,
            textvariable=self.search_var,
            width=25
        )
        self.search_entry.pack(side=LEFT, padx=(0, 5))
        self.search_entry.bind("<Return>", lambda e: self._do_search())
        
        # 搜索列多选
        ttk.Label(search_frame, text="列:").pack(side=LEFT, padx=(5, 2))
        self.column_selector = MultiSelectListbox(
            search_frame,
            self.available_columns,
            on_change=self._on_column_changed
        )
        self.column_selector.pack(side=LEFT, padx=(0, 5))
        
        # 搜索全部文本勾选框
        self.full_text_check = ttk.Checkbutton(
            search_frame,
            text="搜索全文",
            variable=self.search_full_text_var,
            bootstyle="info-round-toggle"
        )
        self.full_text_check.pack(side=LEFT, padx=(5, 5))
        
        # 定位按钮（搜索单个）
        ttk.Button(
            search_frame,
            text="定位",
            command=self._do_search,
            bootstyle="primary",
            width=6
        ).pack(side=LEFT, padx=(0, 2))
        
        # 搜索全部按钮
        ttk.Button(
            search_frame,
            text="搜索全部",
            command=self._do_search_all,
            bootstyle="primary-outline",
            width=8
        ).pack(side=LEFT, padx=(0, 2))
        
        # 上一个/下一个按钮
        ttk.Button(
            search_frame,
            text="◀ 上一个",
            command=self._do_prev,
            bootstyle="secondary-outline",
            width=8
        ).pack(side=LEFT, padx=(0, 2))
        
        ttk.Button(
            search_frame,
            text="下一个 ▶",
            command=self._do_next,
            bootstyle="secondary-outline",
            width=8
        ).pack(side=LEFT, padx=(0, 2))
        
        # 清除按钮
        ttk.Button(
            search_frame,
            text="清除",
            command=self._clear_search,
            bootstyle="secondary-outline",
            width=6
        ).pack(side=LEFT)
        
    
    def update_columns(self, columns: List[str]):
        """更新可搜索的列列表"""
        self.available_columns = ["全部列"] + columns
        self.column_selector.update_values(self.available_columns)
    
    def _on_column_changed(self, selected: List[str]):
        """列选择改变回调"""
        pass
    
    def _do_search(self):
        """执行搜索（定位单个）"""
        keyword = self.search_var.get().strip()
        columns = self.column_selector.get_selected()
        full_text = self.search_full_text_var.get()
        
        if self.on_search_callback:
            self.on_search_callback(keyword, columns, full_text)
    
    def _do_search_all(self):
        """执行搜索全部"""
        keyword = self.search_var.get().strip()
        columns = self.column_selector.get_selected()
        full_text = self.search_full_text_var.get()
        
        if self.on_search_all_callback:
            self.on_search_all_callback(keyword, columns, full_text)
    
    def _do_prev(self):
        """上一个"""
        if self.on_prev_callback:
            self.on_prev_callback()
    
    def _do_next(self):
        """下一个"""
        if self.on_next_callback:
            self.on_next_callback()
    
    def _clear_search(self):
        """清除搜索"""
        self.search_var.set("")
        if self.on_search_callback:
            self.on_search_callback("", ["全部列"], False)
    
    def _do_export(self, export_type: str):
        """执行导出"""
        if self.on_export_callback:
            self.on_export_callback(export_type)
    
    def get_search_params(self) -> Dict:
        """获取当前搜索参数"""
        return {
            "keyword": self.search_var.get().strip(),
            "columns": self.column_selector.get_selected(),
            "full_text": self.search_full_text_var.get()
        }


class DynamicSearchResultsTable(ttk.Frame):
    """动态搜索结果表格 - 根据源表动态创建列"""
    
    def __init__(self, parent, app):
        super().__init__(parent)
        self.app = app
        self.results = []
        self.current_columns = []
        self.sort_column = None
        self.sort_reverse = False
        self._create_ui()
    
    def _create_ui(self):
        """创建UI"""
        # 标题栏
        title_frame = ttk.Frame(self)
        title_frame.pack(fill=X, pady=(0, 5))
        
        self.title_label = ttk.Label(
            title_frame,
            text="🔍 搜索结果 (0条)",
            font=("", 10, "bold")
        )
        self.title_label.pack(side=LEFT)
        
        # 清空按钮
        ttk.Button(
            title_frame,
            text="✕ 清空",
            command=self.clear_results,
            bootstyle="secondary-outline",
            width=8
        ).pack(side=RIGHT)
        
        # 表格容器
        self.table_container = ttk.Frame(self)
        self.table_container.pack(fill=BOTH, expand=YES)
        
        # 初始化空表格
        self._create_table([("info", "提示", 400, False)])
    
    def _create_table(self, columns: List[tuple]):
        """创建表格"""
        for widget in self.table_container.winfo_children():
            widget.destroy()
        
        self.current_columns = columns
        self.sort_column = None
        self.sort_reverse = False
        
        scroll_y = ttk.Scrollbar(self.table_container, orient=VERTICAL)
        scroll_y.pack(side=RIGHT, fill=Y)
        
        scroll_x = ttk.Scrollbar(self.table_container, orient=HORIZONTAL)
        scroll_x.pack(side=BOTTOM, fill=X)
        
        col_ids = [col[0] for col in columns if len(col) >= 1]
        
        self.tree = ttk.Treeview(
            self.table_container,
            columns=col_ids,
            show="headings",
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set,
            height=10
        )
        self.tree.pack(side=LEFT, fill=BOTH, expand=YES)
        
        scroll_y.config(command=self.tree.yview)
        scroll_x.config(command=self.tree.xview)
        
        for col in columns:
            col_id = col[0]
            col_title = col[1] if len(col) > 1 else col_id
            col_width = col[2] if len(col) > 2 else 100
            is_numeric = col[3] if len(col) > 3 else False
            
            # 绑定排序功能
            self.tree.heading(
                col_id, 
                text=col_title,
                command=lambda c=col_id, n=is_numeric: self.sort_by_column(c, n)
            )
            self.tree.column(col_id, width=col_width)
        
        self.tree.bind("<Double-1>", self._on_double_click)
    
    def sort_by_column(self, col, is_numeric):
        """
        点击表头排序
        
        Args:
            col: 列标识
            is_numeric: 是否为数值列
        """
        # 判断是否需要反转排序
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False
        
        # 获取所有项
        items = [(self.tree.set(child, col), child) for child in self.tree.get_children('')]
        
        if not items:
            return
        
        # 排序
        if is_numeric:
            # 数值排序
            def extract_number(item):
                value = item[0]
                try:
                    # 尝试提取数字（处理带单位的情况）
                    import re
                    numbers = re.findall(r'[-+]?\d*\.?\d+', str(value))
                    return float(numbers[0]) if numbers else 0
                except:
                    return 0
            
            items.sort(key=extract_number, reverse=self.sort_reverse)
        else:
            # 文本排序
            items.sort(key=lambda x: str(x[0]).lower(), reverse=self.sort_reverse)
        
        # 重新排列
        for index, (value, child_id) in enumerate(items):
            self.tree.move(child_id, '', index)
        
        # 更新表头显示排序状态
        for col_info in self.current_columns:
            col_id = col_info[0]
            col_title = col_info[1] if len(col_info) > 1 else col_id
            if col_id == col:
                indicator = " ▼" if self.sort_reverse else " ▲"
                self.tree.heading(col_id, text=col_title + indicator)
            else:
                self.tree.heading(col_id, text=col_title)
    
    def show_results(self, results: List[Dict], keyword: str, source_columns: List[tuple]):
        """显示搜索结果"""
        self.results = results
        
        columns = [("_source", "来源", 100, False)]
        for col in source_columns:
            if len(col) == 4:
                columns.append(col)
            elif len(col) == 3:
                columns.append((col[0], col[1], col[2], False))
            elif len(col) == 2:
                columns.append((col[0], col[1], 100, False))
            else:
                columns.append((col[0], str(col[0]), 100, False))
        
        self._create_table(columns)
        
        self.title_label.config(text=f"🔍 搜索结果 ({len(results)}条) - 关键词: {keyword}")
        
        for result in results:
            source = result.get("source", "-")
            values = result.get("values", [])
            item_id = result.get("item_id", "")
            
            row_values = [source] + list(values)
            
            self.tree.insert(
                "",
                "end",
                values=row_values,
                tags=(item_id, source)
            )
    
    def clear_results(self):
        """清空搜索结果"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.title_label.config(text="🔍 搜索结果 (0条)")
        self.results = []
    
    def _on_double_click(self, event):
        """双击定位到原始位置"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = selection[0]
        tags = self.tree.item(item, "tags")
        if tags and len(tags) >= 2:
            item_id = tags[0]
            source = tags[1]
            self.event_generate("<<LocateItem>>", data=f"{source}:{item_id}")


# 保持向后兼容
SearchResultsTable = DynamicSearchResultsTable


def export_table_to_csv(tree: ttk.Treeview, columns: List[tuple], file_path: str):
    """导出表格到CSV文件"""
    with open(file_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        headers = [col[1] if len(col) > 1 else col[0] for col in columns]
        writer.writerow(headers)
        for item in tree.get_children():
            values = tree.item(item, "values")
            writer.writerow(values)


def export_table_to_excel(tree: ttk.Treeview, columns: List[tuple], file_path: str, sheet_name: str = "Sheet1"):
    """导出表格到Excel文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise ImportError("需要安装openpyxl库: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    headers = [col[1] if len(col) > 1 else col[0] for col in columns]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    for row_idx, item in enumerate(tree.get_children(), 2):
        values = tree.item(item, "values")
        for col_idx, value in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    for col_idx, col in enumerate(columns, 1):
        width = col[2] if len(col) > 2 else 100
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(width // 7, 10)
    
    wb.save(file_path)


def export_all_tables_to_excel(tables_data: Dict[str, Dict], file_path: str):
    """导出所有表格到Excel文件（多工作表）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise ImportError("需要安装openpyxl库: pip install openpyxl")
    
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    
    first_sheet = True
    for table_name, table_info in tables_data.items():
        tree = table_info.get("tree")
        columns = table_info.get("columns", [])
        
        if not tree:
            continue
        
        if first_sheet:
            ws = default_sheet
            ws.title = table_name[:31]
            first_sheet = False
        else:
            ws = wb.create_sheet(title=table_name[:31])
        
        headers = [col[1] if len(col) > 1 else col[0] for col in columns]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for row_idx, item in enumerate(tree.get_children(), 2):
            values = tree.item(item, "values")
            for col_idx, value in enumerate(values, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        for col_idx, col in enumerate(columns, 1):
            width = col[2] if len(col) > 2 else 100
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(width // 7, 10)
    
    wb.save(file_path)
